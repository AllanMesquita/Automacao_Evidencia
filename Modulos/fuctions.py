def atualizar(tblPA, aba_tblPA, path, file_name, df_mastersaf, v17, logging, error_log_registro):

    ### IMPORTS
    from Modulos import validacao, atualizar_v2
    import openpyxl as xl
    from openpyxl.styles import PatternFill
    import os
    from datetime import datetime
    import psycopg2

    ###
    global id_arquivo
    # qtd_linhas_tblPA = aba_tblPA.UsedRange.Rows.Count

    nome_evidencia = file_name.split('_')

    # aba_tblPA.Range(f'A{qtd_linhas_tblPA + 1}').Value = nome_evidencia[0] + '_' + nome_evidencia[1]
    # aba_tblPA.Range(f'B{qtd_linhas_tblPA + 1}').Value = 'UpdatePlanEstoque'
    # aba_tblPA.Range(f'C{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
    # aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = 'EmProcessamento'

    # tblPA.Save()

    # qtd_linhas_tblPA = len(aba_tblPA['A'])
    #
    # aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = file_name.split('_')[0]
    # aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'UpdatePlanEstoque'
    # aba_tblPA[f'C{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
    # aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = 'EmProcessamento'

    # tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

    # QUERY ARQUIVO - BD
    try:
        con = psycopg2.connect(
            host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
            dbname="logistic-control",
            user="logisticpsqladmin@psql-itlatam-logisticcontrol",
            password="EsjHSrS69295NzHu342ap6P!N",
            sslmode="require"
        )
        cur = con.cursor()
        cur.execute(
            f'INSERT INTO material_management.mm_tbl_processamento_automacoes (id_tbl, query, '
            f'processamento_inicio, status) VALUES (%s, %s, %s, %s)',
            (
                nome_evidencia[0] + '_' + nome_evidencia[1],
                'UpdatePlanEstoque',
                datetime.now(),
                'EmProcessamento'
            )
        )
        con.commit()
        cur.execute(
            f"SELECT id FROM material_management.mm_tbl_processamento_automacoes WHERE "
            f"id_tbl = '{nome_evidencia[0] + '_' + nome_evidencia[1]}'"
        )
        retorno = cur.fetchall()
        for c in retorno:
            id_arquivo = c[0]
        cur.close()
        con.close()
    except Exception as error:
        logging.basicConfig(filename=error_log_registro, filemode='w', format='%(asctime)s %(message)s')
        logging.critical(f'- {error}', exc_info=True)

        tblPA = xl.open(
            "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")
        tblPA.active
        tblPA_sheet = tblPA.sheetnames
        aba_tblPA = tblPA[tblPA_sheet[0]]

        qtd_linhas_tblPA = len(aba_tblPA['A'])

        aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = nome_evidencia[0] + '_' + nome_evidencia[1]
        aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'UpdatePlanEstoque'
        aba_tblPA[f'C{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
        aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = 'EmProcessamento'

        tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

    finally:
        # print(file_name)
        wb = xl.open(path + file_name)
        wb.active
        sheet = wb.sheetnames
        aba = wb[sheet[0]]

        verif_evid = str(aba["A2"].value)  # Verificador do tipo de evidência

        # Variáveis

        type_evidencia = ""
        qtd_linhas = 0
        resultado = ''
        validacao_local = False

        while bool(aba[f'A{qtd_linhas + 1}'].value) is True:
            qtd_linhas += 1

        if bool(verif_evid) is False:
            print("Erro! Célula vazia")
            # os.replace(path + file_name,
            #            "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\"
            #            "100 BcoDados\\003 Evidencias\\03 Erro\\" + file_name)
            os.replace(path + file_name,
                       "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias"
                       "\\03 Erro\\" + file_name)
        elif len(verif_evid) == 44:
            type_evidencia = "Recebimento"
        # verificador de repetição de caracteres
        elif len(verif_evid) == 24 and "E" in verif_evid:
            type_evidencia = "Expedição"
        elif len(verif_evid) != 44:
            print("Erro! Diferente de 44.")
            # os.replace(path + file_name,
            #            "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\"
            #            "100 BcoDados\\003 Evidencias\\03 Erro\\" + file_name)
            os.replace(path + file_name,
                       "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias"
                       "\\03 Erro\\" + file_name)
        if type_evidencia == 'Recebimento':
            local = aba['G2'].value
            if local.strip() == 'TERCA VIX' or local.strip() == 'AGS RIO' or local.strip() == 'NEXUS SAO':
                pass
            else:
                aba['G2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                validacao_local = True
                resultado = 'Erro local'
        elif type_evidencia == 'Expedição':
            local = aba['D2'].value
            if local.strip() == 'TERCA VIX' or local.strip() == 'AGS RIO' or local.strip() == 'NEXUS SAO':
                pass
            else:
                aba['D2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                validacao_local = True
                resultado = 'Erro local'

        # inicio = now
        # print(now)

        if validacao_local is False:
            if type_evidencia == "Recebimento":
                resultado = validacao.rec_validation(aba, qtd_linhas, nome_evidencia[0] + '_' + nome_evidencia[1])
            if type_evidencia == "Expedição":
                resultado = validacao.exp_validacao(aba, qtd_linhas, nome_evidencia[0] + '_' + nome_evidencia[1])

        wb.save(path + file_name)

        if resultado == 'Erro nos dados' or resultado == 'Erro local':
            os.replace(path + file_name,
                       "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\"
                       "003 Evidencias\\03 Erro\\" + file_name)
                        # Verificar o caso se haver o mesmo arquivo no destino

        elif resultado == 'Sucesso':
            atualizar_v2.popular_V17(aba, qtd_linhas, type_evidencia, df_mastersaf, v17)
            os.replace(path + file_name,
                       "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\"
                       "003 Evidencias\\04 Fluig\\" + file_name)

        # aba_tblPA.Range(f'D{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
        # aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = resultado
        #
        # tblPA.Save()

        # aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
        # aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = resultado
        #
        # tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

        # STATUS FIM ARQUIVO
        try:
            con = psycopg2.connect(
                host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
                dbname="logistic-control",
                user="logisticpsqladmin@psql-itlatam-logisticcontrol",
                password="EsjHSrS69295NzHu342ap6P!N",
                sslmode="require"
            )
            cur = con.cursor()
            cur.execute(
                f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                f"processamento_fim = '{datetime.now()}', "
                f"status = '{resultado}' "
                f"WHERE id = '{id_arquivo}'"
            )
            con.commit()
            cur.close()
            con.close()
        except Exception as error:
            logging.basicConfig(filename=error_log_registro, filemode='w', format='%(asctime)s %(message)s')
            logging.critical(f'- {error}', exc_info=True)

            tblPA = xl.open(
                "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")
            tblPA.active
            tblPA_sheet = tblPA.sheetnames
            aba_tblPA = tblPA[tblPA_sheet[0]]

            qtd_linhas_tblPA = len(aba_tblPA['A'])

            aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = nome_evidencia[0] + '_' + nome_evidencia[1]
            aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'UpdatePlanEstoque'
            aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
            aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = resultado

            tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")
        finally:
            return resultado
