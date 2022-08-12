from turtle import st


def popular_V17(aba, qtd_linhas, type_evid, df_mastersaf, v17):

    import openpyxl as xl
    import pandas as pd
    from openpyxl.styles import PatternFill, Font
    from datetime import datetime
    #import xlwings

    tempo_popular = datetime.now()

    """
        Passar estes carregamento de planilhas para o 'main'.
        Talvez não funcione, pois quando tem alguma alteração, sem salvar, os novos dados não são visualizados.
    """
    #v17 = xl.open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Teste - Gestão Estoque RFID - Estoque Consolidado V17 - 21.03.2022.xlsm", keep_vba=True)
    #v17.active
    v17_sheets = v17.sheetnames
    aba_v17 = v17[v17_sheets[2]]

    #v17 = xlwings.Book("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Teste - Gestão Estoque RFID - Estoque Consolidado V17 - 21.03.2022.xlsm")
    #aba_v17 = xlwings.sheets['ItensArmazenados']

    tbl_rec = xl.open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\tblEvidenciaRecebimento.xlsm", keep_vba=True)
    tbl_rec.active
    tbl_rec_sheets = tbl_rec.sheetnames
    aba_tblRec = tbl_rec[tbl_rec_sheets[0]]

    tbl_exp = xl.open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\tblEvidenciaExpedicao.xlsm", keep_vba=True)
    tbl_exp.active
    tbl_exp_sheets = tbl_exp.sheetnames
    aba_tblExp = tbl_exp[tbl_exp_sheets[1]]

    #df_mastersaf = pd.read_excel("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\INDICADORES\\Bases\\2022 - NFs Entrada Mastersaf.xlsx", sheet_name='NFsEntradaItens')


    colunas_v17 = ['A', 'B', 'H', 'K', 'I', 'J', 'C', 'E', 'F', 'G', 'D']
    colunas_tblRec = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
    colunas_evid = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'J', 'K']
    colunas_tblExp_inserir = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    colunas_tblExp = ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA',
                      'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI']
    colunas_v17_base = ['F', 'L', 'M', 'N', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB',
                        'AC', 'AD', 'AE', 'AF', 'AG', 'AH']
    chaveRelac_dic = {}
    chaveRelac_v17 = {}
    delete_row = []
    lista_dados_evid = []

    ultima_linha_tblRec = len(aba_tblRec['A']) + 1
    ultima_linha_v17 = len(aba_v17['A']) + 1
    #ultima_linha_V17 = aba_v17.range('A2').end('down').row + 1
    ultima_linha_tblExp = len(aba_tblExp['A']) + 1
    backup_ultima_linha_v17 = ultima_linha_v17
    backup_ultima_linha_tblRec = ultima_linha_tblRec

    linha = 2
    linha_exp = 2
    qtd_linhsExp = len(aba_tblExp['H'])
    qtd_linhasV17 = len(aba_v17['G'])
    #qtd_linhaV17 = aba_v17.range('G2').end('down').end

    coluna_v17 = 0
    coluna_v17_base = 0
    coluna_tblRec = 0
    coluna_evid = 0
    coluna_tblExp = 0
    posicao = 0

    chave = ''
    pn = ''
    find_chave = ''

    data = datetime.today()
    data = data.strftime('%d/%m/%Y')

    font = Font(name='Arial', size=9)

    # Carregamento da coluna ChaveRelacionamento da Tbl Expedição

    while linha_exp != qtd_linhsExp + 1:
        chaveRelac_dic[aba_tblExp[f'H{linha_exp}'].value] = linha_exp
        linha_exp += 1

    # print(f'Tempo de carregamento dos arquivos: {datetime.now() - tempo_popular}')
    # print('Início da atualização - Tbl.Recebimento')
    tempo_tblrec = datetime.now()

    if type_evid == 'Recebimento':

        ### INSERÇÃO DA EVIDÊNCIA NA TBL RECEBIMENTO

        while linha != qtd_linhas + 1:
            for col in colunas_tblRec:
                lista_dados_evid.append(aba[f'{col}{linha}'].value)
            for col in colunas_tblRec:
                if col == "H":
                    if type(lista_dados_evid[posicao]) == datetime: 
                        aba_tblRec[f'{col}{ultima_linha_tblRec}'] = datetime.strftime(lista_dados_evid[posicao], '%d/%m/%Y')
                    else:
                        aba_tblRec[f'{col}{ultima_linha_tblRec}'] = lista_dados_evid[posicao]
                else:
                    aba_tblRec[f'{col}{ultima_linha_tblRec}'] = lista_dados_evid[posicao]
                posicao += 1
            linha += 1
            posicao = 0
            ultima_linha_tblRec += 1
            lista_dados_evid.clear()

        # for col in colunas_tblRec:
        #     while linha != qtd_linhas + 1:
        #         aba_tblRec[f'{col}{ultima_linha_tblRec}'] = aba[f'{col}{linha}'].value
        #         linha += 1
        #         ultima_linha_tblRec += 1
        #     linha = 2
        #     ultima_linha_tblRec = backup_ultima_linha_tblRec
        # while linha != qtd_linhas + 1:
            # cell_range = aba[f'A{linha}':f'M{linha}']
            # for cell in cell_range:
            #     for data in cell:
            #         aba_tblRec[f'{colunas_tblRec[coluna_tblRec]}{ultima_linha_tblRec}'] = data.value
            #         coluna_tblRec += 1
            coluna_tblRec = 0
            # linha += 1
            # ultima_linha_tblRec += 1
        #print(f'Tempo de atualização da Tbl.Rec: {datetime.now() - tempo_tblrec}')
        
        linha = 2

        #print('Fim da atualização - Tbl.Recebimento')



        ### INSERÇÃO DA EVIDÊNCIA NA PLANILHA ESTOQUE

        temp = datetime.now()
        v17.active = v17[v17_sheets[2]]
        while linha != qtd_linhas + 1:
            """
                Verificação se a evidência encontra-se na Tbl Expedição.
                Estando na Tbl de Expedição a linha será ignorada.
            """
            if aba[f'K{linha}'].value in chaveRelac_dic:
                if type(aba[f'H{linha}'].value) == datetime:
                    if aba_tblExp[f'E{chaveRelac_dic[aba[f"K{linha}"].value]}'].value >= aba[f'H{linha}'].value:
                        linha += 1
                        continue
                elif type(aba[f'H{linha}'].value) == str:
                    aba_to_date = datetime.strptime(aba[f'H{linha}'].value, '%d/%m/%Y')
                    if type(aba_tblExp[f'E{chaveRelac_dic[aba[f"K{linha}"].value]}'].value) == str:
                        tblExp_to_date = datetime.strptime(aba_tblExp[f'E{chaveRelac_dic[aba[f"K{linha}"].value]}'].value, '%d/%m/%Y')
                        if tblExp_to_date >= aba_to_date:
                            linha += 1
                            continue
                    elif type(aba_tblExp[f'E{chaveRelac_dic[aba[f"K{linha}"].value]}'].value) == datetime:
                        if aba_tblExp[f'E{chaveRelac_dic[aba[f"K{linha}"].value]}'].value >= aba_to_date:
                            linha += 1
                            continue
                    
                # if aba_tblExp[f'E{chaveRelac_dic[aba[f"K{linha}"].value]}'].value >= aba[f'H{linha}'].value:
                #     linha += 1
                #     continue
            """
                Etapa de inserção dos dados na planilha de estoque.
            """
            for col in colunas_v17:
                if col == 'D':
                    aba_v17[f'{col}{ultima_linha_v17}'] = 'Recebimento'
                    aba_v17[f'{col}{ultima_linha_v17}'].font = font
                    #aba_v17.range(f'{col}{ultima_linha_V17}').value = 'Recebimento'
                # elif col == 'U':
                #     aba_v17[f'{col}{ultima_linha_v17}'] = '=@IF(AND([@DtEntrada]="";[@DtEntradaRetorno]="";[@DataEvidencia]="");"";IFS([@DtEntradaRetorno]<>"";TODAY()-[@DtEntradaRetorno];[@DtEntrada]<>"";TODAY()-[@DtEntrada];[@DataEvidencia]<>"";TODAY()-[@DataEvidencia]))'
                #     aba_v17[f'{col}{ultima_linha_v17}'].font = font
                # elif col == 'V':
                #     aba_v17[f'{col}{ultima_linha_v17}'] = '=@IFS([@Dias]="";"";[@Dias]<=30;"1 - 0 a 30 dias";[@Dias]<=60;"2 - 31 a 60 dias";[@Dias]<=90;"3 - 61 a 90 dias";[@Dias]<=180;"4 - 91 a 180 dias";[@Dias]<=365;"5 - 181 a 365 dias";[@Dias]>365;"6 - Acima de 365 dias")'
                #     aba_v17[f'{col}{ultima_linha_v17}'].font = font
                # elif col == 'AH':
                #     aba_v17[f'{col}{ultima_linha_v17}'] = '=@IFS(AND(OR([@OrgEstoque]=22;[@OrgEstoque]=24;[@OrgEstoque]=1;[@OrgEstoque]=28);OR([@Classificacao]="ATIVO";[@Classificacao]="ESTOQUE");[@[PartNumber_Oracle]]<>"";[@[OrdemVenda_Lcto]]<>"");"Ajuste c/ Reserva";AND(OR([@OrgEstoque]=22;[@OrgEstoque]=24;[@OrgEstoque]=1;[@OrgEstoque]=28);OR([@Classificacao]="ATIVO";[@Classificacao]="ESTOQUE");[@[PartNumber_Oracle]]<>"");"Ajuste p/ Estoque";AND(OR([@OrgEstoque]=22;[@OrgEstoque]=24;[@OrgEstoque]=1;[@OrgEstoque]=28);OR([@Classificacao]="ATIVO";[@Classificacao]="ESTOQUE"));"Ajuste c/ Possível Dif. no PN";AND(OR([@OrgEstoque]=22;[@OrgEstoque]=24;[@OrgEstoque]=1;[@OrgEstoque]=28));"Ajuste c/ Risco de Classificação";[@OrgEstoque]="";"Sem Rastreabilidade")'
                #     aba_v17[f'{col}{ultima_linha_v17}'].font = font
                # elif col == 'AT':
                #     aba_v17[f'{col}{ultima_linha_v17}'] = '=IF([@ReservaData]="";"";IF(DAYS(TODAY();[@ReservaData])>30;"Revisar Solicitação";"Ativo"))'
                #     aba_v17[f'{col}{ultima_linha_v17}'].font = font
                else:
                    cell_range = aba[f'{colunas_evid[coluna_evid]}{linha}'].value
                    if colunas_evid[coluna_evid] == 'B':
                        aba_v17[f'{col}{ultima_linha_v17}'] = int(cell_range)
                        aba_v17[f'{col}{ultima_linha_v17}'].font = font
                        #aba_v17.range(f'{col}{ultima_linha_V17}').value = int(cell_range)
                    elif colunas_evid[coluna_evid] == 'H':
                        if type(cell_range) == datetime:
                            aba_v17[f'{col}{ultima_linha_v17}'] = datetime.strftime(cell_range, '%d/%m/%Y')
                        else:
                            aba_v17[f'{col}{ultima_linha_v17}'] = cell_range
                        aba_v17[f'{col}{ultima_linha_v17}'].font = font
                    else:
                        aba_v17[f'{col}{ultima_linha_v17}'] = cell_range
                        #aba_v17.range(f'{col}{ultima_linha_V17}').value = cell_range
                        aba_v17[f'{col}{ultima_linha_v17}'].font = font
                    if colunas_evid[coluna_evid] == 'A':
                        chave = cell_range
                        find_chave = df_mastersaf.loc[df_mastersaf['Chave de Acesso'] == chave]
                        if find_chave.empty:
                            aba_v17[f'{col}{ultima_linha_v17}'].fill = PatternFill(fill_type='solid', fgColor='FF0000')
                            #aba_v17.range(f'{col}{ultima_linha_V17}').color = '#FF0000'
                        else:
                            find_org = find_chave['CNPJ/CPF do Destinatário']
                            if find_org.at[find_org.index[0]] == int('00447484000111'):
                                aba_v17[f'P{ultima_linha_v17}'] = 1
                                aba_v17[f'P{ultima_linha_v17}'].font = font
                                #aba_v17.range(f'P{ultima_linha_V17}').value = 1
                            elif find_org.at[find_org.index[0]] == int('00447484000200'):
                                aba_v17[f'P{ultima_linha_v17}'] = 2
                                aba_v17[f'P{ultima_linha_v17}'].font = font
                                #aba_v17.range(f'P{ultima_linha_V17}').value = 2
                            elif find_org.at[find_org.index[0]] == int('00447484000626'):
                                aba_v17[f'P{ultima_linha_v17}'] = 6
                                aba_v17[f'P{ultima_linha_v17}'].font = font
                                #aba_v17.range(f'P{ultima_linha_V17}').value = 6
                            elif find_org.at[find_org.index[0]] == int('05437734000156'):
                                aba_v17[f'P{ultima_linha_v17}'] = 22
                                aba_v17[f'P{ultima_linha_v17}'].font = font
                                #aba_v17.range(f'P{ultima_linha_V17}').value = 22
                            elif find_org.at[find_org.index[0]] == int('05437734000318'):
                                aba_v17[f'P{ultima_linha_v17}'] = 24
                                aba_v17[f'P{ultima_linha_v17}'].font = font
                                #aba_v17.range(f'P{ultima_linha_V17}').value = 24
                            elif find_org.at[find_org.index[0]] == int('05437734000407'):
                                aba_v17[f'P{ultima_linha_v17}'] = 26
                                aba_v17[f'P{ultima_linha_v17}'].font = font
                                #aba_v17.range(f'P{ultima_linha_V17}').value = 26
                            elif find_org.at[find_org.index[0]] == int('05437734000580'):
                                aba_v17[f'P{ultima_linha_v17}'] = 28
                                aba_v17[f'P{ultima_linha_v17}'].font = font
                                #aba_v17.range(f'P{ultima_linha_V17}').value = 28
                            elif find_org.at[find_org.index[0]] == int('05437734000660'):
                                aba_v17[f'P{ultima_linha_v17}'] = 30
                                aba_v17[f'P{ultima_linha_v17}'].font = font
                                #aba_v17.range(f'P{ultima_linha_V17}').value = 30
                            elif find_org.at[find_org.index[0]] == int('31546914000186'):
                                aba_v17[f'P{ultima_linha_v17}'] = 50
                                aba_v17[f'P{ultima_linha_v17}'].font = font
                                #aba_v17.range(f'P{ultima_linha_V17}').value = 50
                    elif colunas_evid[coluna_evid] == 'D':
                        pn = cell_range
                        find_chave = df_mastersaf.loc[df_mastersaf['Chave de Acesso'] == chave]
                        if find_chave.empty:
                            pass
                        else:
                            find_pn = find_chave.loc[find_chave['Cód. Produto'] == pn]
                            if find_pn.empty:
                                aba_v17[f'{col}{ultima_linha_v17}'].fill = PatternFill(fill_type='solid', fgColor='FF0000')
                                #aba_v17.range(f'{col}{ultima_linha_V17}').color = '#FF0000'
                            else:
                                find_valor = find_pn['Valor Unitário Comercial']
                                valor = find_valor.at[find_valor.index[0]]   #.replace(',', '.')
                                aba_v17[f'AE{ultima_linha_v17}'] = valor     #float(valor)
                                #aba_v17.range(f'AE{ultima_linha_V17}').value = float(valor)
                                aba_v17[f'AE{ultima_linha_v17}'].font = font
                coluna_evid += 1
            linha += 1
            ultima_linha_v17 += 1
            coluna_evid = 0

        #print(f'Tempo para inserir na V17: {datetime.now() - temp}')

        ### AJUSTE DA TABELA

        #aba_v17.tables['ItensArmazenados'].ref = f'A2:AT{len(aba_v17["A"])}'
        v17.active = v17[v17_sheets[2]]
        v17.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Backup V17\\Backup V17.1\\Gestão Estoque RFID - Estoque Consolidado V17.1 - 05.05.2022.xlsm")
        #v17.save()
        tbl_exp.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\tblEvidenciaExpedicao.xlsm")
        tbl_rec.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\tblEvidenciaRecebimento.xlsm")
        #print(f'Tempo popular: {datetime.now() - tempo_popular}')

    elif type_evid == 'Expedição':

        ### INSERÇÃO DA EVIDÊNCIA NA TBL EXPEDIÇÃO

        tempo_inserir = datetime.now()
        # print(f'Inicio do processo de retirada dos itens.')

        v17.active = v17[v17_sheets[2]]

        linha = 2
        posicao = 0

        while linha != qtd_linhas + 1:
            for col in colunas_tblExp_inserir:
                lista_dados_evid.append(aba[f'{col}{linha}'].value)
            for col in colunas_tblExp_inserir:
                if col == "E":
                    if type(lista_dados_evid[posicao]) == datetime: 
                        aba_tblExp[f'{col}{ultima_linha_tblExp}'] = datetime.strftime(lista_dados_evid[posicao], '%d/%m/%Y')
                    else:
                        aba_tblExp[f'{col}{ultima_linha_tblExp}'] = lista_dados_evid[posicao]
                else:
                    aba_tblExp[f'{col}{ultima_linha_tblExp}'] = lista_dados_evid[posicao]
                posicao += 1
            linha += 1
            posicao = 0
            ultima_linha_tblExp += 1
            lista_dados_evid.clear()

        #print(f'Tempo para inserir na Tbl Expedição: {datetime.now() - tempo_inserir}')
        #tbl_exp.save("C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\tblEvidenciaExpedicao.xlsm")
        """
        O delete_rows funciona no 'xl.open', então precisa somente abrir o arquivo apenas uma vez, no 'main' talvez.
        E após a exclusão necessita ajustar a tabale, pois quando se deleta a linha não deleta da tabela.
        """
        #v17 = xl.open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\"
        #              "Teste - Gestão Estoque RFID - Estoque Consolidado V17 - 14.01.2022 - Copia (2).xlsm", keep_vba=True)
        #v17.active
        #v17_sheets = v17.sheetnames
        #aba_v17 = v17[v17_sheets[2]]

        ultima_linha_v17 = len(aba_v17["A"]) + 1
        #ultima_linha_V17 = aba_v17.range('A2').end('down').row + 1
        linha = 2
        linha_v17 = 3
        qtd_linhsExp = len(aba_tblExp["A"]) + 1
        #tblExp_to_date = ""

        while linha_exp != qtd_linhsExp + 1:
            chaveRelac_dic[aba_tblExp[f'H{linha_exp}'].value] = linha_exp
            linha_exp += 1

        tempo_exp = datetime.now()

        while linha != qtd_linhas + 1:
            cell_range = aba[f'H{linha}'].value
            # print(f'Linha evidência: {linha}.')
            while linha_v17 != ultima_linha_v17:
                # print(f'Linha V17: {linha_v17}.')
                if cell_range == aba_v17[f'G{linha_v17}'].value:
                #if cell_range == aba_v17.range(f'G{linha_v17}').value:
                    
                    ### VERIFICAÇÃO DO FORMATO DA DATA NA EVIDÊNCIA
                    if type(aba[f'E{linha}'].value) == datetime:
                        aba_to_date = aba[f'E{linha}'].value
                    else:
                        aba_to_date = datetime.strptime(aba[f'E{linha}'].value, '%d/%m/%Y')

                    ### VERIFICAÇÃO DO FORMATO DA DATA NA TBL EXPEDIÇÃO
                    if type(aba_tblExp[f'E{chaveRelac_dic[cell_range]}'].value) == str:
                        tblExp_to_date = datetime.strptime(aba_tblExp[f'E{chaveRelac_dic[cell_range]}'].value, '%d/%m/%Y')
                    elif type(aba_tblExp[f'E{chaveRelac_dic[cell_range]}'].value) == datetime:
                        tblExp_to_date = aba_tblExp[f'E{chaveRelac_dic[cell_range]}'].value
                    
                    if tblExp_to_date >= aba_to_date:
                        for col in colunas_tblExp:
                            if col == 'AH':
                                aba_tblExp[f'AH{chaveRelac_dic[cell_range]}'] = 'data'
                            elif col == 'AI':
                                aba_tblExp[f'AI{chaveRelac_dic[cell_range]}'] = 'Automatizado'
                            else:
                                aba_tblExp[f'{col}{chaveRelac_dic[cell_range]}'] = aba_v17[f'{colunas_v17_base[coluna_v17_base]}{linha_v17}'].value
                                # aba_tblExp[f'{col}{chaveRelac_dic[cell_range]}'] = aba_v17.range(f'{colunas_v17_base[coluna_v17_base]}{linha_v17}').value
                            coluna_v17_base += 1
                        aba_v17.delete_rows(linha_v17, 1)
                        # aba_v17.range(f'{linha_v17}:{linha_v17}').api.Delete()
                        break
                    else:
                        break
                linha_v17 += 1
                coluna_v17_base = 0
            linha_v17 = 3
            coluna_v17_base = 0
            linha += 1

        chaveRelac_dic.clear()

        # AJUSTE DA TABELA

        # aba_v17.tables['ItensArmazenados'].ref = f'A2:AT{len(aba_v17["A"])}'

        # SALVAR PLANILHAS

        # print(f'Tempo de verificação e exclusão da V17: {datetime.now() - tempo_exp}')
        v17.active = v17[v17_sheets[2]]
        v17.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Backup V17\\"
                 "Backup V17.1\\Gestão Estoque RFID - Estoque Consolidado V17.1 - 05.05.2022.xlsm")
        # v17.save()
        tbl_exp.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\"
                     "tblEvidenciaExpedicao.xlsm")
        