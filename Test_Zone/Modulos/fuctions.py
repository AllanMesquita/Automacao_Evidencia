def atualizar(tblPA, aba_tblPA, path, file_name, v17):

    ### IMPORTS
    from Modulos import validacao, atualizar_v2
    import openpyxl as xl
    from openpyxl.styles import PatternFill
    import os
    from datetime import datetime
    import psycopg2

    ###
    global id_arquivo
    qtd_linhas_tblPA = len(aba_tblPA['A'])

    nome_evidencia = file_name.split('_')

    aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = nome_evidencia[0] + '_' + nome_evidencia[1]
    aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'UpdatePlanEstoque'
    aba_tblPA[f'C{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
    aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = 'EmProcessamento'

    tblPA.save(
        "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\"
        "06 Lixeira\\Testes\\02 Tabela\\tblProcessamentoAutomacoes.xlsx")

    # QUERY ARQUIVO - BD
    # con = psycopg2.connect(
    #     host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
    #     dbname="logistic-control",
    #     user="logisticpsqladmin@psql-itlatam-logisticcontrol",
    #     password="EsjHSrS69295NzHu342ap6P!N",
    #     sslmode="require"
    # )
    # cur = con.cursor()
    # cur.execute(
    #             f'INSERT INTO material_management.mm_tbl_processamento_automacoes (id_tbl, query, '
    #             f'processamento_inicio, status) VALUES (%s, %s, %s, %s)',
    #            (
    #             nome_evidencia[0] + '_' + nome_evidencia[1],
    #             'UpdatePlanEstoque',
    #             datetime.now(),
    #             'EmProcessamento'
    #            )
    #            )
    # con.commit()
    # cur.execute(
    #             f"SELECT id FROM material_management.mm_tbl_processamento_automacoes WHERE "
    #             f"id_tbl = '{nome_evidencia[0] + '_' + nome_evidencia[1]}'"
    #            )
    # retorno = cur.fetchall()
    # for c in retorno:
    #     id_arquivo = c[0]
    # cur.close()
    # con.close()

    # print(file_name)
    wb = xl.open(path + file_name)
    wb.active
    sheet = wb.sheetnames
    aba = wb[sheet[0]]

    verif_evid = str(aba["A2"].value)  # Verificador do tipo de evidência

    # Variáveis

    type_evidencia = ""
    qtd_linhas = 0
    resultado = ''
    validacao_local = False

    while bool(aba[f'A{qtd_linhas + 1}'].value) is True:
        qtd_linhas += 1

    if bool(verif_evid) is False:
        print("Erro! Célula vazia")
        os.replace(path + file_name,
                   "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\"
                   "100 BcoDados\\003 Evidencias\\03 Erro\\" + file_name)
    elif len(verif_evid) == 44:
        type_evidencia = "Recebimento"
    # verificador de repetição de caracteres
    elif len(verif_evid) == 24 and "E" in verif_evid:
        type_evidencia = "Expedição"
    elif len(verif_evid) != 44:
        print("Erro! Diferente de 44.")
        os.replace(path + file_name,
                   "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\"
                   "100 BcoDados\\003 Evidencias\\03 Erro\\" + file_name)

    if type_evidencia == 'Recebimento':
        if aba['G2'].value == 'TERCA VIX' or aba['G2'].value == 'JR SAO' or aba['G2'].value == 'JR RIO' or aba['G2'].value == 'AGS RIO' or aba['G2'].value == 'NEXUS SAO':
            pass
        else:
            aba['G2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            validacao_local = True
            resultado = 'Erro local'
    elif type_evidencia == 'Expedição':
        if aba['D2'].value == 'TERCA VIX' or aba['D2'].value == 'JR SAO' or aba['D2'].value == 'JR RIO' or aba['D2'].value == 'AGS RIO' or aba['D2'].value == 'NEXUS SAO':
            pass
        else:
            aba['D2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            validacao_local = True
            resultado = 'Erro local'

    # inicio = now
    # print(now)

    if validacao_local is False:
        if type_evidencia == "Recebimento":
            resultado = validacao.rec_validation(aba, qtd_linhas, file_name)
        if type_evidencia == "Expedição":
            resultado = validacao.exp_validacao(aba, qtd_linhas, file_name)

    wb.save(path + file_name)

    if resultado == 'Erro nos dados' or resultado == 'Erro local':
        os.replace(path + file_name,
                   "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\"
                   "003 Evidencias\\06 Lixeira\\Testes\\03 Erro\\" + file_name)
                    # Verificar o caso se haver o mesmo arquivo no destino

    elif resultado == 'Sucesso':
        atualizar_v2.popular_V17(aba, qtd_linhas, type_evidencia, v17)
        os.replace(path + file_name,
                   "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\"
                   "003 Evidencias\\06 Lixeira\\Testes\\04 Fluig\\" + file_name)

    aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
    aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = resultado

    tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\"
               "003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\tblProcessamentoAutomacoes.xlsx")

    # STATUS FIM ARQUIVO
    # con = psycopg2.connect(
    #     host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
    #     dbname="logistic-control",
    #     user="logisticpsqladmin@psql-itlatam-logisticcontrol",
    #     password="EsjHSrS69295NzHu342ap6P!N",
    #     sslmode="require"
    # )
    # cur = con.cursor()
    # cur.execute(
    #             f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
    #             f"processamento_fim = '{datetime.now()}', "
    #             f"status = '{resultado}' "
    #             f"WHERE id = '{id_arquivo}'"
    # )
    # con.commit()
    # cur.close()
    # con.close()

    return resultado
