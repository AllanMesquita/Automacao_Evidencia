# Dicionários

# from tkinter.ttk import Style
import dateutil.parser

repeticao_RFID = {}
repeticao_SN = {}
# erros_linhas = []  # Lista com todos os erros encontrados numa única linha

# Variáveis

# global linha_validada
global linha
global error_chave
global error_PO
global error_PN
global error_RFID
global error_SN
global error_Date
global error_ChaveRel


def rec_validation(aba, qtd_linhas, file_name):
    # Imports

    from datetime import datetime
    from openpyxl.styles import PatternFill
    import pandas as pd
    import warnings
    from Modulos.class_erros import Error, SaveError
    from dateutil.parser import parse
    import psycopg2

    # tempo_recebimento = datetime.now()

    global error_chave

    # linha_validada = 0
    linha = 2
    dict_chaves = {}

    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
    # print('carregamento V17')
    dfV17 = pd.read_excel(
        "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\07 Inventario\\Gestão Estoque RFID - Estoque Consolidado V17.1 - 02.12.22.xlsm",
        sheet_name="ItensArmazenados")
    # print('Carregamento tbl recebimento')
    dfTblRec = pd.read_excel(
        "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\tblEvidenciaRecebimento.xlsm")
    dfRFID = dfV17['Unnamed: 8'].tolist()
    dfSerial = dfV17['Unnamed: 9'].tolist()
    dfTblRec_ChaveRelacionamento = dfTblRec['ChaveRelacionamento'].tolist()
    # df_nfEntrada = pd.read_excel(
    #     "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\INDICADORES\\Bases\\2022 á 2027 - Nfs Entrada Mastersaf.xlsx",
    #     sheet_name='Dados dos Itens'
    # )

    print('Início da validação - Recebimento')

    while linha != qtd_linhas + 1:
        # print(linha)

        # celula_error = aba[f'N{linha}'].value
        error = Error()

        error_chave = 0
        error_PO = 0
        error_PN = 0
        error_RFID = 0
        error_SN = 0
        error_Date = 0
        error_ChaveRel = 0

        ### VALIDAÇÃO DA CHAVE DE NOTA FISCAL

        # linha_validada += 1

        cell_range = aba[f"A{linha}"].value

        ### Pesquisa da nota em Banco de Dados
        # Update connection string information
        host = "psql-itlatam-logisticcontrol.postgres.database.azure.com"
        dbname = "logistic-control"
        user = "logisticpsqladmin@psql-itlatam-logisticcontrol"
        password = "EsjHSrS69295NzHu342ap6P!N"
        sslmode = "require"
        # Construct connection string
        conn_string = "host={0} user={1} dbname={2} password={3} sslmode={4}".format(host, user, dbname,
                                                                                     password,
                                                                                     sslmode)
        conn = psycopg2.connect(conn_string)
        print("Connection established")
        cursor = conn.cursor()

        cursor.execute(f"SELECT chave_acesso FROM material_management.master_saf_entrada WHERE chave_acesso = '{cell_range}'")
        pesquisa = cursor.fetchall()

        cursor.close()
        conn.close()

        if bool(pesquisa) is False:
            error.chave_bd()
            error_chave += 1

        if bool(cell_range) is False:
            aba[f"A{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            # aba[f'N{linha}'] = Error.empty()
            error.empty()
            error_chave += 1
        elif type(cell_range) != int and cell_range.isnumeric() is False:
            aba[f"A{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            # aba[f'N{linha}'] = Error(celula_error).chave()
            error.chave()
            error_chave += 1
        elif len(cell_range) != 44:
            aba[f"A{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            # aba[f'N{linha}'] = Error(celula_error).chave()
            error.chave()
            error_chave += 1
        elif len(cell_range) == 44:
            pass
        if cell_range not in dict_chaves:
            dict_chaves[cell_range] = 1
        else:
            dict_chaves[cell_range] += 1

        ### VALIDAÇÃO DO PEDIDO DE COMPRA (PO)

        # linha_validada += 1
        cell_range = aba[f"B{linha}"].value
        if bool(cell_range) is False:
            error_PO += 1
            aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            # aba[f'N{linha}'] = Error(celula_error).empty()
            error.empty()
        elif type(cell_range) != int and cell_range.isnumeric() is False:
            if "K" not in cell_range and "k" not in cell_range:
                error_PO += 1
                aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                # aba[f'N{linha}'] = Error(celula_error).po()
                error.po()
            elif len(cell_range[1:]) > 5 or len(cell_range[1:]) < 5 or cell_range[1:].isnumeric() is False:
                error_PO += 1
                aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                # aba[f'N{linha}'] = Error(celula_error).po()
                error.po()
            else:
                aba[f"B{linha}"] = cell_range[1:]
        elif len(str(cell_range)) > 5 or len(str(cell_range)) < 5:
            error_PO += 1
            aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            # aba[f'N{linha}'] = Error(celula_error).po()
            error.po()
        else:
            pass

        ### VALIDAÇÃO DO PART-NUMBER

        # linha_validada += 1
        cell_range = str(aba[f"D{linha}"].value)
        if bool(cell_range) is False or cell_range is None:
            aba[f"D{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            # aba[f'N{linha}'] = Error(celula_error).empty()
            error.empty()
            error_PN += 1
        elif "!" in cell_range or \
                "@" in cell_range or \
                "$" in cell_range or \
                "%" in cell_range or \
                "&" in cell_range or \
                "*" in cell_range or \
                ")" in cell_range or \
                "'" in cell_range or \
                ":" in cell_range or \
                ";" in cell_range:
            aba[f"D{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            # aba[f'N{linha}'] = Error(celula_error).part_number()
            error.part_number()
            error_PN += 1
        elif "1P" in cell_range:
            aba[f"D{linha}"] = cell_range[2:]
        elif "30P" in cell_range:
            aba[f"D{linha}"] = cell_range[3:]
        else:
            pass

        ### VALIDAÇÃO RFID DO PRODUTO

        # linha_validada += 1

        cell_range = aba[f"E{linha}"].value
        if bool(cell_range) is False:
            aba[f"E{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            # aba[f'N{linha}'] = Error(celula_error).empty()
            error.empty()
            error_RFID += 1
        elif len(str(cell_range)) != 24:
            aba[f"E{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            # aba[f'N{linha}'] = Error(celula_error).rfid()
            error.rfid()
            error_RFID += 1
        elif "E" != cell_range[0]:
            aba[f"E{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            # aba[f'N{linha}'] = Error(celula_error).rfid()
            error.rfid()
            error_RFID += 1
        if cell_range in repeticao_RFID:
            aba[f'E{linha}'].fill = PatternFill(fill_type='solid', fgColor='38B9FF')
            repeticao_RFID[cell_range].fill = PatternFill(fill_type='solid', fgColor='38B9FF')
            # aba[f'N{linha}'] = Error(celula_error).rfid_repetido()
            error.rfid_repetido()
            error_RFID += 1
        else:
            repeticao_RFID[cell_range] = aba[f'E{linha}']
        if cell_range in dfRFID:
            aba[f'E{linha}'].fill = PatternFill(fill_type='solid', fgColor="9B9B9B")
            # aba[f'N{linha}'] = Error(celula_error).rfid_v17()
            error.rfid_v17()
            error_RFID += 1
        else:
            pass

        ### VALIDAÇÃO DO SERIAL NUMBER

        # linha_validada += 1
        cell_range = aba[f"F{linha}"].value
        if bool(cell_range) is False:
            aba[f"F{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            # aba[f'N{linha}'] = Error(celula_error).empty()
            error.empty()
            error_SN += 1
        elif "S" == str(cell_range)[0]:
            aba[f"F{linha}"] = cell_range[1:]
        if "!" in str(cell_range) or \
                "@" in str(cell_range) or \
                "$" in str(cell_range) or \
                "%" in str(cell_range) or \
                "&" in str(cell_range) or \
                "*" in str(cell_range) or \
                "(" in str(cell_range) or \
                ")" in str(cell_range) or \
                "'" in str(cell_range) or \
                ":" in str(cell_range) or \
                "/" in str(cell_range):
            aba[f"F{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            # aba[f'N{linha}'] = Error(celula_error).serial_number()
            error.serial_number()
            error_SN += 1
        if cell_range in repeticao_SN:
            aba[f'F{linha}'].fill = PatternFill(fill_type='solid', fgColor='38B9FF')
            repeticao_SN[cell_range].fill = PatternFill(fill_type='solid', fgColor='38B9FF')
            # aba[f'N{linha}'] = Error(celula_error).serial_number_repetido()
            error.serial_number_repetido()
            error_SN += 1
        else:
            repeticao_SN[cell_range] = aba[f'F{linha}']
        if cell_range in dfSerial:
            aba[f'F{linha}'].fill = PatternFill(fill_type='solid', fgColor='9B9B9B')
            # aba[f'N{linha}'] = Error(celula_error).serial_number_v17()
            error.serial_number_v17()
            error_SN += 1
        else:
            pass

        ### VALIDAÇÃO DA DATA

        ### Formatar a data, para ser copiada para as tabelas e estoque ###

        # linha_validada += 1

        cell_range = aba[f'H{linha}'].value

        # try:
            # parse(cell_range)

        if type(cell_range) == datetime:
            if cell_range > datetime.today():
                aba[f"H{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF0000')
                error.data_maior()
                error_Date += 1
            else:
                pass
        else:
            try:
                data = parse(cell_range).strptime(cell_range, '%d/%m/%Y')
                if data > datetime.today():
                    aba[f"H{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF0000')
                    error.data_maior()
                    error_Date += 1
                else:
                    pass
        # if data.day <= 12:
        #     data = datetime.strptime(datetime.strftime(parse(cell_range), "%m/%d/%Y"), "%d/%m/%Y")
        # # else:
        # #     data = data
        # if data > datetime.today():
        #     aba[f"H{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF0000')
        #     error.data_maior()
        #     error_Date += 1
        # else:
        #     pass
            except Exception as erros:
                print(erros)
                aba[f"H{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF0000')
                error.data()
                error_Date += 1

        # if bool(aba[f'H{linha}'].value) is False or cell_range is None:
        #     aba[f"H{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF7B00')
        #     # aba[f'N{linha}'] = Error(celula_error).empty()
        #     error.empty()
        #     error_Date += 1
        # if bool(cell_range) is True:
        #     if type(aba[f'H{linha}'].value) == datetime:
        #         cell_range = aba[f'H{linha}'].value
        #     else:
        #         cell_range = datetime.strptime(aba[f'H{linha}'].value, '%d/%m/%Y')
        # if type(cell_range) != datetime:
        #     aba[f"H{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF0000')
        #     # aba[f'N{linha}'] = Error(celula_error).data()
        #     error.data()
        #     error_Date += 1
        # elif cell_range > datetime.today():
        #     aba[f"H{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF0000')
        #     # aba[f'N{linha}'] = Error(celula_error).data_maior()
        #     error.data_maior()
        #     error_Date += 1
        # else:
        #     pass

        ### CHAVE DE RELACIONAMENTO

        aba[f'K{linha}'] = str(aba[f'E{linha}'].value) + str(aba[f'G{linha}'].value)

        try:
            cell_range = aba[f'H{linha}'].value
            data = parse(str(cell_range))
            if data.day <= 12:
                data = datetime.strptime(datetime.strftime(parse(str(aba[f'H{linha}'].value)), "%m/%d/%Y"), "%d/%m/%Y")

            # if bool(aba[f'H{linha}'].value) is True:
            #     if type(aba[f'H{linha}'].value) == datetime:
            #         cell_range = aba[f'H{linha}'].value
            #     else:
            #         cell_range = datetime.strptime(aba[f'H{linha}'].value, '%d/%m/%Y')

            if aba[f'K{linha}'].value in dfTblRec_ChaveRelacionamento:
                tem_df = dfTblRec.loc[dfTblRec['ChaveRelacionamento'] == aba[f'K{linha}'].value]
                for i, row in tem_df.iterrows():
                    # print(row['DataEvidencia'], '-', datetime.strptime(aba[f'H{linha}'].value, '%d/%m/%Y'))
                    if type(row['DataEvidencia']) == datetime:
                        data_verif = row['DataEvidencia']
                    else:
                        data_verif = datetime.strptime(row['DataEvidencia'], '%d/%m/%Y')
                    if data_verif >= data:
                        aba[f'K{linha}'].fill = PatternFill(fill_type='solid', fgColor='E7E200')
                        # aba[f'N{linha}'] = Error(celula_error).chave_relacionamento()
                        error.chave_relacionamento()
                        error_ChaveRel += 1
                    else:
                        pass
            else:
                pass
        # else:
        except dateutil.parser.ParserError:
            error_ChaveRel += 1

        ### LANÇAMENTO BANCO DE DADOS - DATA

        date = datetime.today()
        aba[f'L{linha}'] = date.strftime("%d/%m/%Y %H:%M")

        ### LANÇAMENTO BANCO DE DADOS - USUÁRIO

        aba[f'M{linha}'] = 'Automatizado'

        if error_chave > 0 or error_PO > 0 or error_PN > 0 \
                or error_RFID > 0 or error_SN > 0 or error_Date > 0 or error_ChaveRel:
            aba[f'N{linha}'] = error.retornar()
            # print(error.dic_erros)
            # save = SaveError(aba, linha, 'Recebimento', error.dic_erros, file_name)
            # print(save.connect())
            linha += 1
        else:
            linha += 1

    aba['A1'] = 'ChaveNF_Entrada'
    aba['B1'] = 'PedidoCompra'
    aba['C1'] = 'RFID_CxMaster/TagAtivo'
    aba['D1'] = 'PartNumber'
    aba['E1'] = 'RFID_Produto'
    aba['F1'] = 'SerialNumber'
    aba['G1'] = 'Local'
    aba['H1'] = 'DataEvidencia'
    aba['I1'] = 'Usuario(email)'
    aba['J1'] = 'ObsRecebimento'
    aba['K1'] = 'ChaveRelacionamento'
    aba['L1'] = 'LctoBD_Data'
    aba['M1'] = 'LctoBD_Usuario'

    # print(f"Erro ChaveNF:{error_chave:>8}"
    # f"\nErro Pedido: {error_PO:>8}"
    # f"\nErro PartNumber: {error_PN:>4}"
    # f"\nErro RFID: {error_RFID:>10}"
    # f"\nErro SerialNumber:{error_SN:>3}"
    # f"\nErro Data: {error_Date:>10}"
    # f"\nErro Chave Relacionamento: {error_ChaveRel:>10}"
    # f"\nCélulas validadas: {linha_validada:>4}")

    # print(f'Tempo validação recebimento: {datetime.now() - tempo_recebimento}')
    # print('Fim da validação')
    #
    # for chave in dict_chaves:
    #
    #     erro = Error()

        ### Pesquisa substituida de Pandas por SQL - 14.12.20222
        # itens_chave = []
        #
        # temp_df = df_nfEntrada.loc[df_nfEntrada['Unnamed: 17'] == chave]
        # qtd_chave = temp_df['Unnamed: 28']
        #
        # if temp_df.empty:
        #     continue
        # else:
        #     for itens in qtd_chave:
        #         itens_chave.append(float(itens.replace('.', '').replace(',', '.')))

    #     # Update connection string information
    #     host = "psql-itlatam-logisticcontrol.postgres.database.azure.com"
    #     dbname = "logistic-control"
    #     user = "logisticpsqladmin@psql-itlatam-logisticcontrol"
    #     password = "EsjHSrS69295NzHu342ap6P!N"
    #     sslmode = "require"
    #     # Construct connection string
    #     conn_string = "host={0} user={1} dbname={2} password={3} sslmode={4}".format(host, user, dbname,
    #                                                                                  password,
    #                                                                                  sslmode)
    #     conn = psycopg2.connect(conn_string)
    #     print("Connection established")
    #     cursor = conn.cursor()
    #
    #     cursor.execute(
    #         f"SELECT quantidade_com FROM material_management.master_saf_entrada_itens WHERE chave_acesso = '{chave}'")
    #     pesquisa = cursor.fetchall()
    #
    #     cursor.close()
    #     conn.close()
    #
    #     soma_bd = 0
    #
    #     if bool(pesquisa) is False:
    #         aba[f'A{linha}'].fill = PatternFill(fill_type='solid', fgColor='33CC33')
    #         aba[f'O{linha}'] = 'Chave de Nota Fiscal não consta em Banco Dados'
    #         error_chave += 1
    #         continue
    #     else:
    #         for dado in pesquisa:
    #             for item in dado:
    #                 soma_bd += item
    #
    #         # if sum(itens_chave) == dict_chaves[chave]:
    #         if soma_bd == dict_chaves[chave]:
    #             pass
    #         else:
    #             linha = 2
    #             while linha != qtd_linhas + 1:
    #                 cell_range = aba[f'A{linha}'].value
    #                 if cell_range == chave:
    #                     aba[f'A{linha}'].fill = PatternFill(fill_type='solid', fgColor='33CC33')
    #                     aba[f'O{linha}'] = 'Quantidade do RFID diferente da Nota Fiscal'
    #                     error_chave += 1
    #                     erro.quantidade()
    #                     save = SaveError(aba, linha, 'Recebimento', erro.dic_erros, file_name)
    #                     save.connect()
    #                 else:
    #                     linha += 1
    #                     continue
    #                 linha += 1
    #     # itens_chave.clear()
    #
    # repeticao_RFID.clear()
    # repeticao_SN.clear()

    if error_chave > 0 or error_PO > 0 or error_PN > 0 \
            or error_RFID > 0 or error_SN > 0 or error_Date > 0 or error_ChaveRel:
        aba['N1'] = 'ERROS'
        return 'Erro nos dados'
    else:
        return 'Sucesso'


def exp_validacao(aba, qtd_linhas, file_name):
    # Imports
    from openpyxl.styles import PatternFill
    from datetime import datetime
    import pandas as pd
    from Modulos.class_erros import Error, SaveError

    # Variáveis

    tempo_exp = datetime.now()

    # linha_validada = 0
    linha = 2
    # error_chave = 0
    # error_RFID = 0
    # error_Date = 0
    # error_ChaveRel = 0
    dict_chaves = {}

    dfTblExp = pd.read_excel(
        "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\tblEvidenciaExpedicao.xlsm",
        sheet_name='Evidencias')
    dfTblExp_ChaveRelacionamento = dfTblExp['ChaveRelacionamento'].tolist()
    df_NF_saida = pd.read_excel(
        "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\INDICADORES\\Bases\\2022 á 2027 - NFs Saída Mastersaf.xlsx",
        sheet_name='Dados dos Itens')

    # print('Início da validação - Expedição')

    while linha != qtd_linhas + 1:

        # celula_error = aba[f'K{linha}'].value
        error = Error()

        error_chave = 0
        error_RFID = 0
        error_Date = 0
        error_ChaveRel = 0

        ### VALIDAÇÃO RFID DO PRODUTO

        # linha_validada += 1
        # print(linha)
        cell_range = aba[f"A{linha}"].value
        if bool(cell_range) is False:
            aba[f"A{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            # aba[f'K{linha}'] = error.empty()
            error.empty()
            error_RFID += 1
        elif len(str(cell_range)) != 24:
            aba[f"A{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            # aba[f'K{linha}'] = error.rfid()
            error.rfid()
            error_RFID += 1
        elif "E" != cell_range[0]:
            aba[f"A{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            # aba[f'K{linha}'] = error.rfid()
            error.rfid()
            error_RFID += 1
        if cell_range in repeticao_RFID:
            aba[f'A{linha}'].fill = PatternFill(fill_type='solid', fgColor='38B9FF')
            repeticao_RFID[cell_range].fill = PatternFill(fill_type='solid', fgColor='38B9FF')
            # aba[f'K{linha}'] = error.rfid_repetido()
            error.rfid_repetido()
            error_RFID += 1
        else:
            repeticao_RFID[cell_range] = aba[f'A{linha}']

        ### VALIDAÇÃO DA CHAVE DE NOTA FISCAL

        # linha_validada += 1
        cell_range = aba[f"B{linha}"].value

        if bool(cell_range) is False:
            aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            # aba[f'K{linha}'] = error.empty()
            error.empty()
            error_chave += 1
        elif type(cell_range) != int and cell_range.isnumeric() is False:
            aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            # aba[f'K{linha}'] = error.chave()
            error.chave()
            error_chave += 1
        elif len(cell_range) != 44:
            aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            # aba[f'K{linha}'] = error.chave()
            error.chave()
            error_chave += 1
        elif len(cell_range) == 44:
            pass
        if cell_range not in dict_chaves:
            dict_chaves[cell_range] = 1
        else:
            dict_chaves[cell_range] += 1

        ### VALIDAÇÃO DA DATA

        # linha_validada += 1
        cell_range = aba[f"E{linha}"].value
        if bool(cell_range) is False:
            aba[f"E{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF7B00')
            # aba[f'K{linha}'] = error.empty()
            error.empty()
            error_Date += 1
        if bool(cell_range):
            if type(cell_range) == datetime:
                cell_range = aba[f'E{linha}'].value
            elif type(cell_range) == str:
                cell_range = datetime.strptime(cell_range, '%d/%m/%Y')

        if type(cell_range) != datetime:
            aba[f"E{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF0000')
            # aba[f'K{linha}'] = error.data()
            error.data()
            error_Date += 1
        elif cell_range > datetime.today():
            aba[f"E{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF0000')
            # aba[f'K{linha}'] = error.data_maior()
            error.data_maior()
            error_Date += 1
        else:
            pass

        ### CHAVE DE RELACIONAMENTO

        aba[f"H{linha}"] = str(aba[f"A{linha}"].value) + str(aba[f"D{linha}"].value)

        # if aba[f"H{linha}"].value in dfTblExp_ChaveRelacionamento:
        #     tem_df = dfTblExp.loc[dfTblExp['ChaveRelacionamento'] == aba[f"H{linha}"].value]
        #     for i, row in tem_df.iterrows():
        #         if str(row['DataEvidencia']) >= str(aba[f"E{linha}"].value):
        #             aba[f"H{linha}"].fill = PatternFill(fill_type='solid', fgColor='E7E200')
        #             error_ChaveRel += 1
        #         else:
        #             continue
        # else:
        #     pass

        if bool(aba[f'E{linha}'].value) is True:
            if type(aba[f'E{linha}'].value) == datetime:
                cell_range = aba[f'E{linha}'].value
            else:
                cell_range = datetime.strptime(aba[f'E{linha}'].value, '%d/%m/%Y')

            if aba[f'H{linha}'].value in dfTblExp_ChaveRelacionamento:
                tem_df = dfTblExp.loc[dfTblExp['ChaveRelacionamento'] == aba[f'H{linha}'].value]
                for i, row in tem_df.iterrows():
                    if type(row['DataEvidencia']) == datetime:
                        data_verif = row['DataEvidencia']
                    else:
                        data_verif = datetime.strptime(row['DataEvidencia'], '%d/%m/%Y')
                    if data_verif >= cell_range:
                        aba[f'H{linha}'].fill = PatternFill(fill_type='solid', fgColor='E7E200')
                        # aba[f'K{linha}'] = error.chave_relacionamento()
                        error.chave_relacionamento()
                        error_ChaveRel += 1
                    else:
                        pass
            else:
                pass
        else:
            error_ChaveRel += 1

        ### lANÇAMENTO BANCO DE DADOS - DATA

        date = datetime.today()
        aba[f"I{linha}"] = date.strftime("%d/%m/%Y %H:%M")

        ### LANÇAMENTO BANCO DE DADOS - USUÁRIO

        aba[f"J{linha}"] = 'Automatizado'

        if error_RFID > 0 or error_chave > 0 or error_Date > 0 or error_ChaveRel > 0:
            aba[f'K{linha}'] = error.retornar()
            save = SaveError(aba, linha, 'Expedição', error.dic_erros, file_name)
            save.connect()
            linha += 1
        else:
            linha += 1

    aba[f'A1'] = 'RFID_Produto'
    aba[f'B1'] = 'ChaveNF_Saida'
    aba[f'C1'] = 'OrdemVenda'
    aba[f'D1'] = 'Local'
    aba[f'E1'] = 'DataEvidencia'
    aba[f'F1'] = 'Usuario(email)'
    aba[f'G1'] = 'ObsExpedicao'
    aba[f'H1'] = 'ChaveRelacionamento'
    aba[f'I1'] = 'LctoBD_Data'
    aba[f'J1'] = 'LctoBD_Usuario'

    for chaves in dict_chaves:
        itens_chave = []

        temp_df = df_NF_saida.loc[df_NF_saida['Unnamed: 17'] == chaves]
        qtd_chave = temp_df['Unnamed: 28']

        if temp_df.empty:
            continue
        else:
            for itens in qtd_chave:
                itens_chave.append(float(itens.replace(',', '.')))

            if sum(itens_chave) == dict_chaves[chaves]:
                pass
            else:
                linha = 2
                while linha != qtd_linhas + 1:
                    cell_range = aba[f'B{linha}'].value
                    if cell_range == chaves:
                        aba[f'B{linha}'].fill = PatternFill(fill_type='solid', fgColor='33CC33')
                        aba[f'L{linha}'] = 'Quantidade do RFID diferente da Nota Fiscal'
                        erro = Error()
                        erro.quantidade()
                        save = SaveError(aba, linha, 'Expedição', erro.dic_erros, file_name)
                        save.connect()
                    else:
                        linha += 1
                        continue

                    linha += 1
        itens_chave.clear()

    # print(f"Erro ChaveNF:{error_chave:>8}"
    # f"\nErro RFID: {error_RFID:>10}"
    # f"\nErro Data: {error_Date:>10}"
    # f"\nErro Chave Relacionamento: {error_ChaveRel}"
    # f"\nCálulas validadas: {linha_validada:>4}")

    # print(dict_chaves)
    # print(datetime.now() - tempo_exp)
    # print('Fim da validação')
    repeticao_RFID.clear()
    dict_chaves.clear()

    if error_RFID > 0 or error_chave > 0 or error_Date > 0 or error_ChaveRel > 0:
        aba['K1'] = 'ERROS'
        return 'Erro nos dados'
    else:
        return 'Sucesso'
