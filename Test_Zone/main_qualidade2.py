"""
    Main Script
"""

import openpyxl as xl
from datetime import datetime
from Modulos import validacao
from openpyxl.styles import PatternFill
import os
import sys
import win32com.client
import logging
import psycopg2

now = datetime.now()

name_log = str(
    'C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Script\\Error_Log\\main_qualidade\\Error_Log_' + datetime.strftime(
        datetime.today(), '%d-%m-%Y %H.%M') + '.txt')

resultado = ''
id_arquivo = ''
id = ''

# path = "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\" + sys.argv[1] + "\\"
# file_name = sys.argv[2]
# path = "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\Temp\\"
# file_name = "202203281100 35220305437734000156550010000133101100010619_Expedição.xlsx"

path = 'C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\01 Processamento\\'

files = os.listdir(
    "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\01 Processamento\\")
len_pasta = len(files)

win32 = win32com.client.Dispatch('Excel.Application')
win32.Visible = False
tblPA = win32.Workbooks.Open(
    "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\tblProcessamentoAutomacoes.xlsx")
aba_tblPA = tblPA.Worksheets('tbl')

verificar_status = False
qtd_linhas_tblPA = aba_tblPA.UsedRange.Rows.Count

# Inserir Query
aba_tblPA.Range(f'A{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(), '%Y%m%d%H%M')
aba_tblPA.Range(f'B{qtd_linhas_tblPA + 1}').Value = 'QualidadeEvidencia'
aba_tblPA.Range(f'C{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')

query_name = datetime.strftime(datetime.now(), '%Y%m%d%H%M')
query_id = query_name
con = psycopg2.connect(
    host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
    dbname="logistic-control",
    user="logisticpsqladmin@psql-itlatam-logisticcontrol",
    password="EsjHSrS69295NzHu342ap6P!N",
    sslmode="require"
)
cur = con.cursor()
# Pesquisa 'EmProcessamento'
cur.execute(f"SELECT status FROM material_management.mm_tbl_processamento_automacoes WHERE status = 'EmProcessamento'")
retorno = cur.fetchall()
if bool(retorno) is False:
    cur.execute(f'INSERT INTO material_management.mm_tbl_processamento_automacoes (id_tbl, query, processamento_inicio,'
                f'status) VALUES (%s, %s, %s, %s)',
                (
                 query_id,
                 'QualidadeEvidencia',
                 datetime.now(),
                 'EmProcessamento'
                 )
                )
    con.commit()
    cur.execute(f"SELECT id FROM material_management.mm_tbl_processamento_automacoes WHERE id_tbl = '{query_id}'")
    retorno = cur.fetchall()
    for c in retorno:
        id = c[0]
else:
    cur.execute(f"INSERT INTO material_management.mm_tbl_processamento_automacoes (id_tbl, query, processamento_inicio,"
                f" processamento_fim, status) VALUES(%s, %s, %s, %s, %s)",
                (
                 query_id,
                 'QualidadeEvidencia',
                 datetime.now(),
                 datetime.now(),
                 'Error - EmProcessamento'
                )
                )
    con.commit()
    cur.close()
    con.close()

for value in range(1, qtd_linhas_tblPA + 1):
    if str(aba_tblPA.Range(f'E{value}')) == "EmProcessamento":
        verificar_status = True
        continue

if verificar_status is False:
    var_linha = qtd_linhas_tblPA + 1
    aba_tblPA.Range(f'C{var_linha}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
    aba_tblPA.Range(f'E{var_linha}').Value = 'EmProcessamento'
    tblPA.Save()

    try:
        if len_pasta > 0:

            for file_name in files:

                qtd_linhas_tblPA = aba_tblPA.UsedRange.Rows.Count

                nome_evidencia = file_name.split('_')

                # aba_tblPA.Range(f'A{qtd_linhas_tblPA + 1}').Value = file_name.split('_')[0]
                aba_tblPA.Range(f'A{qtd_linhas_tblPA + 1}').Value = nome_evidencia[0] + '_' + nome_evidencia[1]
                aba_tblPA.Range(f'B{qtd_linhas_tblPA + 1}').Value = 'QualidadeEvidencia'
                aba_tblPA.Range(f'C{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = 'EmProcessamento'

                tblPA.Save()

                # QUERY ARQUIVO
                cur.execute(
                            f'INSERT INTO material_management.mm_tbl_processamento_automacoes (id_tbl, query, '
                            f'processamento_inicio, status) VALUES (%s, %s, %s, %s)',
                           (
                            nome_evidencia[0] + '_' + nome_evidencia[1],
                            'QualidadeEvidencia',
                            datetime.now(),
                            'EmProcessamento'
                           )
                           )
                con.commit()
                cur.execute(
                            f"SELECT id FROM material_management.mm_tbl_processamento_automacoes WHERE "
                            f"id_tbl = '{nome_evidencia[0] + '_' + nome_evidencia[1]}'"
                           )
                retorno = cur.fetchall()
                for c in retorno:
                    id_arquivo = c[0]

                wb = xl.load_workbook(path + file_name)
                wb.active
                sheet = wb.sheetnames
                aba = wb[sheet[0]]

                verif_evid = str(aba["A2"].value)  # Verificador do tipo de evidência

                # Variáveis

                type_evidencia = ""
                validacao_local = False
                qtd_linhas = len(aba["A"])
                resultado = ''

                if bool(verif_evid) is False:
                    # print("Erro!")
                    aba['A2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                    resultado = 'Erro nos dados'
                elif len(verif_evid) == 44:
                    type_evidencia = "Recebimento"
                    # verificador de repetição de caracteres
                    chave_nf = str(verif_evid)
                    for c in chave_nf:
                        if chave_nf.count(c) == 44:
                            aba['A2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                            resultado = 'Erro nos dados'
                            break
                        else:
                            continue
                elif len(verif_evid) == 24 and "E" in verif_evid:
                    type_evidencia = "Expedição"
                elif len(verif_evid) != 44:
                    # print("Erro!")
                    aba['A2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                    resultado = 'Erro nos dados'

                if type_evidencia == 'Recebimento':
                    if aba['G2'].value == 'TERCA VIX' or aba['G2'].value == 'JR SAO' or aba['G2'].value == 'JR RIO' or \
                            aba['G2'].value == 'AGS RIO' or aba['G2'].value == 'NEXUS SAO':
                        pass
                    else:
                        aba['G2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                        validacao_local = True
                        resultado = 'Erro local'
                elif type_evidencia == 'Expedição':
                    if aba['D2'].value == 'TERCA VIX' or aba['D2'].value == 'JR SAO' or aba['D2'].value == 'JR RIO' or \
                            aba['D2'].value == 'AGS RIO' or aba['D2'].value == 'NEXUS SAO':
                        pass
                    else:
                        aba['D2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                        validacao_local = True
                        resultado = 'Erro local'

                inicio = now
                # print(now)

                if validacao_local == False:
                    if type_evidencia == "Recebimento":
                        resultado = validacao.rec_validation(aba, qtd_linhas,
                                                             nome_evidencia[0] + '_' + nome_evidencia[1])
                    if type_evidencia == "Expedição":
                        resultado = validacao.exp_validacao(aba, qtd_linhas,
                                                            nome_evidencia[0] + '_' + nome_evidencia[1])

                wb.save(path + file_name)

                if resultado == 'Erro nos dados' or resultado == 'Erro local':
                    os.replace(path + file_name,
                               "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\03 Erro\\" + file_name)
                    # os.replace(path + file_name, "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\03 Erro\\" + file_name)
                    # Verificar o caso se haver o mesmo arquivo no destino
                # elif resultado == 'Sucesso':
                # atualizar.popular_V17(aba, qtd_linhas)

                # if error_chave > 0:
                #     print("Houve erro na coluna Chave NF")
                # else:
                #     print("Coluna Chave NF validada")

                aba_tblPA.Range(f'D{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = resultado

                tblPA.Save()

                # STATUS FIM ARQUIVO
                cur.execute(
                            f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                            f"processamento_fim = '{datetime.now()}',"
                            f"status = '{resultado}' "
                            f"WHERE id = '{id_arquivo}'")
                con.commit()

            aba_tblPA.Range(f'D{var_linha}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
            aba_tblPA.Range(f'E{var_linha}').Value = 'Sucesso'
            tblPA.Save()
            win32.Application.Quit()

            # STATUS FIM PROCESSO
            cur.execute(
                        f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                        f"processamento_fim = '{datetime.now()}', "
                        f"status = 'Sucesso' "
                        f"WHERE id = '{id}'")
            con.commit()
            cur.close()
            con.close()
        else:
            resultado = 'Sem arquivos na pasta.'  ### Bloco movido para a linha 148 - 05/05/2022
            # print('teste')
            aba_tblPA.Range(f'D{var_linha}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
            aba_tblPA.Range(f'E{var_linha}').Value = resultado

            # STATUS SEM ARQUIVO NA PASTA
            cur.execute(
                        f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                        f"processamento_fim = '{datetime.now()}', "
                        f"status = '{resultado}' "
                        f"WHERE id = '{id}'")
            con.commit()
            cur.close()
            con.close()

    except Exception as error:
        logging.basicConfig(filename=name_log, filemode='w', format='%(asctime)s %(message)s')
        logging.critical(f'- {error}', exc_info=True)

        aba_tblPA.Range(f'D{var_linha}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
        aba_tblPA.Range(f'E{var_linha}').Value = 'Error-log'

        aba_tblPA.Range(f'D{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(),
                                                                              '%d/%m/%Y %H:%M')  ### Bloco movido para a linha 152 - 05/05/2022
        aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = 'Error-log'
        tblPA.Save()
        win32.Application.Quit()

        # STATUS ERROR-LOG
        '''Update query_id'''
        cur.execute(
                    f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                    f"processamento_fim = '{datetime.now()}',"
                    f"status = 'Error-log' "
                    f"WHERE id = '{id}'")
        con.commit()
        '''Update id_arquivo'''
        cur.execute(
                    f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                    f"processamento_fim = '{datetime.now()}',"
                    f"status = 'Error-log' "
                    f"WHERE id = '{id_arquivo}'")
        con.commit()
        cur.close()
        con.close()

else:
    aba_tblPA.Range(f'D{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
    aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = 'Error - EmProcessamento'
    tblPA.Save()
    win32.Application.Quit()

    outlook = win32com.client.Dispatch("outlook.application")

    mail = outlook.CreateItem(0)

    mail.To = 'allan.mesquita@global.ntt'
    mail.Subject = 'Erro-Log - main_qualidade.py - "EmProcessamento"'
    mail.HTMLBody = '<h3>This is HTML Body</h3>'
    mail.Body = f"""Houve um erro na varificação de qualidade das evidências.
Outro Processo encontra-se com o status "EmProcessamento".

Att.

Python"""

    mail.Send()

    # except Exception as error:
    #     logging.basicConfig(filename=name_log, filemode='w', format='%(asctime)s %(message)s')
    #     logging.critical(f'- {error}', exc_info=True)
    #
    #     aba_tblPA.Range(f'D{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')   ### Bloco movido para a linha 152 - 05/05/2022
    #     aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = 'Error-log'
    #     tblPA.Save()
    #     win32.Application.Quit()

    # else:
    #     resultado = 'Sem arquivos na pasta.'      ### Bloco movido para a linha 148 - 05/05/2022
    #     # print('teste')

    # except Exception as error:
    #     logging.basicConfig(filename=name_log, filemode='w', format='%(asctime)s %(message)s')
    #     logging.critical(f'- {error}', exc_info=True)

# print(inicio)
print(resultado)
print(f'Tempo total: {datetime.now() - now}')
