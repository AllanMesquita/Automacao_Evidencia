"""
    Main Script
"""
# from curses import qiflush
import logging
from datetime import datetime
import win32com.client
import openpyxl as xl
import psycopg2

import Modulos.fuctions

global id
global id_arquivo
name_log = str(
    'C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Script\\Error_Log\\Error_Log_' + datetime.strftime(
        datetime.today(), '%d-%m-%Y %H.%M') + '.txt')

# win32 = win32com.client.Dispatch("Excel.Application")
# win32.Visible = False
# tblPA = win32.Workbooks.Open("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")
# aba_tblPA = tblPA.Worksheets('tbl')
tblPA = xl.open(
    "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\tblProcessamentoAutomacoes.xlsx")
tblPA.active
tblPA_sheets = tblPA.sheetnames
aba_tblPA = tblPA[tblPA_sheets[0]]

linha_tblPA = 2
# qtd_linhas_tblPA = aba_tblPA.UsedRange.Rows.Count
qtd_linhas_tblPA = len(aba_tblPA['A'])
verificar_status = False
local_emprocessamento = ""

# aba_tblPA.Range(f'A{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(), '%Y%m%d%H%M')
# aba_tblPA.Range(f'B{qtd_linhas_tblPA + 1}').Value = 'UpdatePlanEstoque'
# aba_tblPA.Range(f'C{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.today(), '%Y%m%d%H%M')
aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'UpdatePlanEstoque'
aba_tblPA[f'C{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')

# QUERY NAME - BD
query_name = datetime.strftime(datetime.now(), '%Y%m%d%H%M')
query_id = query_name

# print(qtd_linhas_tblPA)

for value in range(1, qtd_linhas_tblPA + 1):
    # print(value, aba_tblPA.Range(f'E{value}'), type(aba_tblPA.Range(f'E{value}')))
    # if str(aba_tblPA.Range(f'E{value}')) == "EmProcessamento":
    if str(aba_tblPA[f'E{value}'].value) == 'EmProcessamento':
        verificar_status = True
        # local_emprocessamento = str(aba_tblPA.Range(f'E{value}'))
        local_emprocessamento = str(aba_tblPA[f'E{value}'])
        continue
# continue

# Pesquisa 'EmProcessamento'
con = psycopg2.connect(
    host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
    dbname="logistic-control",
    user="logisticpsqladmin@psql-itlatam-logisticcontrol",
    password="EsjHSrS69295NzHu342ap6P!N",
    sslmode="require"
)
cur = con.cursor()
cur.execute(f"SELECT status FROM material_management.mm_tbl_processamento_automacoes WHERE status = 'EmProcessamento'")
retorno = cur.fetchall()
if bool(retorno) is False:
    cur.execute(f'INSERT INTO material_management.mm_tbl_processamento_automacoes (id_tbl, query, processamento_inicio,'
                f'status) VALUES (%s, %s, %s, %s)',
                (
                    query_id,
                    'UpdatePlanEstoque',
                    datetime.now(),
                    'EmProcessamento'
                )
                )
    con.commit()
    cur.execute(f"SELECT id FROM material_management.mm_tbl_processamento_automacoes WHERE id_tbl = '{query_id}'")
    retorno = cur.fetchall()
    for c in retorno:
        id = c[0]
    cur.close()
    con.close()
else:
    verificar_status = True

# print(verificar_status)	

if verificar_status is False:
    var_linha = qtd_linhas_tblPA + 1
    # aba_tblPA.Range(f'C{var_linha}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
    # aba_tblPA.Range(f'E{var_linha}').Value = 'EmProcessamento'
    aba_tblPA[f'C{var_linha}'] = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
    aba_tblPA[f'E{var_linha}'] = 'EmProcessamento'
    # tblPA.Save()
    tblPA.save(
        "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\tblProcessamentoAutomacoes.xlsx")

    try:
        import openpyxl as xl
        from openpyxl.styles import PatternFill
        # import tkinter
        # from tkinter import filedialog
        import os
        import pandas as pd
        from Modulos import validacao, atualizar_v2, fuctions

        # root = tkinter.Tk()
        #
        # root.filename = tkinter.filedialog.askopenfilename(initialdir="/Users/Mesqu", filetypes=(("Excel files", "*.xlsx"),
        #                                                                                           ("Excel files", "*.xlsm")))
        #
        # local = root.filename

        now = datetime.now()

        # files = os.listdir("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\01 Processamento\\")
        files = os.listdir(
            "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\01 Processamento\\")

        if len(files) > 0:

            lista_recebimento = []
            lista_expedicao = []
            resultado = ''

            # path = "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\01 Processamento\\"
            path = "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\01 Processamento\\"
            # file_name = "Arquivo_Teste.xlsx"

            # v17 = xl.open("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\001 Estoque\\Gestão Estoque RFID - Estoque Consolidado V17.1.xlsm", keep_vba=True)
            v17 = xl.open(
                "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Backup V17\\Backup V17.1\\Gestão Estoque RFID - Estoque Consolidado V17.1 - 05.05.2022.xlsm",
                keep_vba=True)
            # v17.active
            # v17_sheets = v17.sheetnames
            # aba_v17 = v17[v17_sheets[2]]

            df_mastersaf = pd.read_excel(
                "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\INDICADORES\\Bases\\2022 - NFs Entrada Mastersaf.xlsx",
                sheet_name='NFsEntradaItens')

            # win32 = win32com.client.Dispatch('Excel.Application')
            # win32.Visible = False
            # tblrec = win32.Workbooks.Open("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\Controle_Status_V17.xlsx")
            # aba_tblrec = tblrec.Worksheets('Sheet1')

            # planilha_status = xl.open("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\Controle_Status_V17.xlsx")
            # planilha_status.active
            # planilha_status_sheet = planilha_status.sheetnames
            # aba_planilha = planilha_status[planilha_status_sheet[0]]
            #
            # linha_planilha = len(aba_planilha['A'])
            # aba_planilha[f'A2'] = 'Inicio'
            # aba_planilha[f'B2'] = datetime.now()

            # planilha_status.save("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\Controle_Status_V17.xlsx")

            # aba_tblrec.Range('A2').Value = 'Inicio'
            # aba_tblrec.Range('B2').Value = datetime.now()
            #
            # tblrec.Save()

            for file in files:
                if 'Recebimento' in file:
                    lista_recebimento.append(file)
                elif 'Expedição' in file:
                    lista_expedicao.append(file)

            if len(lista_recebimento) > 0:
                for file_name in lista_recebimento:
                    resultado = Modulos.fuctions.atualizar(tblPA, aba_tblPA, path, file_name, df_mastersaf, v17)

            if len(lista_expedicao) > 0:
                for file_name in lista_expedicao:
                    resultado = fuctions.atualizar(tblPA, aba_tblPA, path, file_name, df_mastersaf, v17)

                '''
                    Seguência abaixo alocada na função 'Modulos.fuctions.atualizar()'.
                '''
                '''
                # qtd_linhas_tblPA = aba_tblPA.UsedRange.Rows.Count
                qtd_linhas_tblPA = len(aba_tblPA['A'])

                # aba_tblPA.Range(f'A{qtd_linhas_tblPA + 1}').Value = file_name.split('_')[0]
                # aba_tblPA.Range(f'B{qtd_linhas_tblPA + 1}').Value = 'UpdatePlanEstoque'
                # aba_tblPA.Range(f'C{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                # aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = 'EmProcessamento'
                aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = file_name.split('_')[0]
                aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'UpdatePlanEstoque'
                aba_tblPA[f'C{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = 'EmProcessamento'

                # tblPA.Save()
                tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\tblProcessamentoAutomacoes.xlsx")

                # print(file_name)
                wb = xl.open(path + file_name)
                wb.active
                sheet = wb.sheetnames
                aba = wb[sheet[0]]

                verif_evid = str(aba["A2"].value)  # Verificador do tipo de evidência

                # Variáveis

                type_evidencia = ""
                #temp_val = ""
                qtd_linhas = 0
                resultado = ''
                validacao_local = False

                while bool(aba[f'A{qtd_linhas + 1}'].value) is True:
                    qtd_linhas += 1

                if bool(verif_evid) is False:
                    print("Erro! Célula vazia")
                    os.replace(path + file_name, "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\03 Erro\\" + file_name)
                elif len(verif_evid) == 44:
                    type_evidencia = "Recebimento"
                    # verificador de repetição de caracteres
                elif len(verif_evid) == 24 and "E" in verif_evid:
                    type_evidencia = "Expedição"
                elif len(verif_evid) != 44:
                    print("Erro! Diferente de 44.")
                    os.replace(path + file_name, "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\03 Erro\\" + file_name)

                if type_evidencia == 'Recebimento':
                    if aba['G2'].value == 'TERCA VIX' or aba['G2'].value == 'JR SAO' or aba['G2'].value == 'JR RIO' or aba['G2'].value == 'AGS RIO' or aba['G2'].value == 'NEXUS SAO':
                        pass
                    else:
                        aba['G2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                        validacao_local = True
                        resultado = 'Erro local'
                elif type_evidencia == 'Expedição':
                    if aba['D2'].value == 'TERCA VIX' or aba['D2'].value == 'JR SAO' or aba['D2'].value == 'JR RIO' or aba['D2'].value == 'AGS RIO' or aba['D2'].value == 'NEXUS SAO':
                        pass
                    else:
                        aba['D2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                        validacao_local = True
                        resultado = 'Erro local'

                #inicio = now
                #print(now)

                if validacao_local is False:
                    if type_evidencia == "Recebimento":
                        resultado = validacao.rec_validation(aba, qtd_linhas)
                    if type_evidencia == "Expedição":
                        resultado = validacao.exp_validacao(aba, qtd_linhas)

                wb.save(path + file_name)

                if resultado == 'Erro nos dados' or resultado == 'Erro local':
                    # os.replace(path + file_name, "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\03 Erro\\" + file_name)
                    os.replace(path + file_name, "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\03 Erro\\" + file_name)
                    # Verificar o caso se haver o mesmo arquivo no destino
                elif resultado == 'Sucesso':
                    atualizar_v2.popular_V17(aba, qtd_linhas, type_evidencia, df_mastersaf, v17)
                    # os.replace(path + file_name, "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\04 Fluig\\" + file_name)
                    os.replace(path + file_name, "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\04 Fluig\\" + file_name)

                # if error_chave > 0:
                #     print("Houve erro na coluna Chave NF")
                # else:
                #     print("Coluna Chave NF validada")

                # linha_planilha = len(aba_planilha['A'])

                # aba_tblPA.Range(f'D{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                # aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = resultado
                aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = resultado

                # tblPA.Save()
                tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\tblProcessamentoAutomacoes.xlsx")
                '''

            # aba_planilha[f'A2'] = 'Fim'
            # aba_planilha[f'B2'] = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
            # aba_planilha[f'C2'] = datetime.now() - now
            #
            # planilha_status.save("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\Controle_Status_V17.xlsx")

            # aba_tblrec.Range('A2').Value = 'Fim'
            #
            # tblrec.Save()

            # aba_tblPA.Range(f'D{var_linha}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
            # aba_tblPA.Range(f'E{var_linha}').Value = 'Sucesso'
            # tblPA.Save()
            # win32.Application.Quit()
            aba_tblPA[f'D{var_linha}'] = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
            aba_tblPA[f'E{var_linha}'] = 'Sucesso'

            tblPA.save(
                "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\tblProcessamentoAutomacoes.xlsx")

            # STATUS FIM PROCESSO
            con = psycopg2.connect(
                host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
                dbname="logistic-control",
                user="logisticpsqladmin@psql-itlatam-logisticcontrol",
                password="EsjHSrS69295NzHu342ap6P!N",
                sslmode="require"
            )
            cur = con.cursor()
            cur.execute(
                        f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                        f"processamento_fim = '{datetime.now()}', "
                        f"status = 'Sucesso' "
                        f"WHERE id = '{id}'"
            )
            con.commit()
            cur.close()
            con.close()

        else:
            resultado = "Sem arquivos na pasta '01 Processamento'."
            # aba_tblPA.Range(f'D{var_linha}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%H')
            # aba_tblPA.Range(f'E{var_linha}').Value = resultado
            aba_tblPA[f'D{var_linha}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
            aba_tblPA[f'E{var_linha}'] = resultado

            # STATUS SEM ARQUIVO NA PASTA
            con = psycopg2.connect(
                host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
                dbname="logistic-control",
                user="logisticpsqladmin@psql-itlatam-logisticcontrol",
                password="EsjHSrS69295NzHu342ap6P!N",
                sslmode="require"
            )
            cur = con.cursor()
            cur.execute(
                        f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                        f"processamento_fim = '{datetime.now()}', "
                        f"status = '{resultado}' "
                        f"WHERE id = '{id}'")
            con.commit()
            cur.close()
            con.close()

        print(resultado)
        print(f'Tempo total: {datetime.now() - now}')

    except Exception as error:
        logging.basicConfig(filename=name_log, filemode='w', format='%(asctime)s %(message)s')
        logging.critical(f'- {error}', exc_info=True)

        # v17.save("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\001 Estoque\\Gestão Estoque RFID - Estoque Consolidado V17.1.xlsm")
        v17.save(
            "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Backup V17\\Backup V17.1\\Gestão Estoque RFID - Estoque Consolidado V17.1 - 05.05.2022.xlsm")

        ### STATUS ERROR V17
        # aba_tblPA.Range(f'D{var_linha}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
        # aba_tblPA.Range(f'E{var_linha}').Value = 'Error-log'
        aba_tblPA[f'D{var_linha}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
        aba_tblPA[f'E{var_linha}'] = 'Error-log'

        ### STATUS ERROR EVIDÊNCIA
        # aba_tblPA.Range(f'D{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
        # aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = 'Error-log'
        # tblPA.Save()
        # win32.Application.Quit()
        aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
        aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = 'Error-log'
        tblPA.save(
            "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\tblProcessamentoAutomacoes.xlsx")

        # STATUS ERROR-LOG
        con = psycopg2.connect(
            host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
            dbname="logistic-control",
            user="logisticpsqladmin@psql-itlatam-logisticcontrol",
            password="EsjHSrS69295NzHu342ap6P!N",
            sslmode="require"
        )
        cur = con.cursor()
        '''Update query_id'''
        cur.execute(
                    f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                    f"processamento_fim = '{datetime.now()}',"
                    f"status = 'Error-log' "
                    f"WHERE id = '{id}'")
        con.commit()
        '''Update id_arquivo'''
        cur.execute(
                    f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                    f"processamento_fim = '{datetime.now()}',"
                    f"status = 'Error-log' "
                    f"WHERE status = 'EmProcessamento'")
        con.commit()
        cur.close()
        con.close()

else:
    # aba_tblPA.Range(f'D{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
    # aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = 'Error - EmProcessamento'
    # tblPA.Save()
    # win32.Application.Quit()
    aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
    aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = 'Error - EmProcessamento'
    tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\tblProcessamentoAutomacoes.xlsx")

    con = psycopg2.connect(
        host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
        dbname="logistic-control",
        user="logisticpsqladmin@psql-itlatam-logisticcontrol",
        password="EsjHSrS69295NzHu342ap6P!N",
        sslmode="require"
    )
    cur = con.cursor()
    cur.execute(f"INSERT INTO material_management.mm_tbl_processamento_automacoes (id_tbl, query, processamento_inicio,"
                f" processamento_fim, status) VALUES(%s, %s, %s, %s, %s)",
                (
                    query_id,
                    'UpdatePlanEstoque',
                    datetime.now(),
                    datetime.now(),
                    'Error - EmProcessamento'
                )
                )
    con.commit()
    cur.close()
    con.close()

    '''
        ENVIO DE E-MAIL
    '''
    outlook = win32com.client.Dispatch("outlook.application")

    mail = outlook.CreateItem(0)

    mail.To = 'allan.mesquita@global.ntt'
    mail.Subject = 'Erro-Log - main_v2.py - "EmProcessamento"'
    mail.HTMLBody = '<h3>This is HTML Body</h3>'
    mail.Body = f"""Houve um erro na atualização da Planilha de estoque.
    O processo {local_emprocessamento} encontra-se com o status "EmProcessamento".
    
    Att.
    
    Python"""

    mail.Send()

    print('Erro - EmProcessamento')

# print(datetime.now())
# print(repeticao_RFID)
# print(Modulos.rec_validation())
# print(error_PO)
# print(error_PN)
# print(error_RFID)
# print(error_SN)
# print(error_Date)
# print(f"{linha_validada:,}")
