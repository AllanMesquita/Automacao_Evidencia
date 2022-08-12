# Dicionários

repeticao_RFID = {}
repeticao_SN = {}

# Variáveis

global linha_validada
global linha
global error_chave
global error_PO
global error_PN
global error_RFID
global error_SN
global error_Date
# global error_usuario
global error_ChaveRel


def rec_validation(aba, qtd_linhas):

    # Imports

    from datetime import datetime
    from openpyxl.styles import PatternFill
    import pandas as pd
    import warnings

    tempo_recebimento = datetime.now()

    linha_validada = 0
    linha = 2
    error_chave = 0
    error_PO = 0
    error_PN = 0
    error_RFID = 0
    error_SN = 0
    error_Date = 0
    # error_usuario = 0
    error_ChaveRel = 0

    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

    dfV17 = pd.read_excel(
        "C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\"
        "Teste - Gestão Estoque RFID - Estoque Consolidado V17 - 14.01.2022 - Copia.xlsm", sheet_name="ItensArmazenados")
    dfTblRec = pd.read_excel(
        "C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\tblEvidenciaRecebimento.xlsm")
    dfRFID = dfV17['Unnamed: 8'].tolist()
    dfSerial = dfV17['Unnamed: 9'].tolist()
    dfTblRec_ChaveRelacionamento = dfTblRec['ChaveRelacionamento'].tolist()
    # list_V17 = df['RFID_Produto'].tolist()

    print('Início da validação - Recebimento')

    while linha != qtd_linhas + 1:

        ### VALIDAÇÃO DA CHAVE DE NOTA FISCAL

        # print(linha)
        linha_validada += 1
        cell_range = aba[f"A{linha}"].value
        if bool(cell_range) is False:
            aba[f"A{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            error_chave += 1
            # linha += 1
            # continue
        elif type(cell_range) != int and cell_range.isnumeric() is False:
            aba[f"A{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            error_chave += 1
            # linha += 1
            # continue
        elif len(cell_range) != 44:
            aba[f"A{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            error_chave += 1
            # linha += 1
            # continue
        elif len(cell_range) == 44:
            # linha += 1
            # continue
            pass

        # linha = 2

        ### VALIDAÇÃO DO PEDIDO DE COMPRA (PO)

        # while linha != qtd_linhas + 1:
        # print(linha)
        linha_validada += 1
        cell_range = aba[f"B{linha}"].value
        if bool(cell_range) is False:
            error_PO += 1
            aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            # linha += 1
            # continue
        elif type(cell_range) != int and cell_range.isnumeric() is False:
            if "K" not in cell_range and "k" not in cell_range:
                error_PO += 1
                aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                # linha += 1
                # continue
            elif len(cell_range[1:]) > 5 or cell_range[1:].isnumeric() is False:
                error_PO += 1
                aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                # linha += 1
                # continue
            else:
                aba[f"B{linha}"] = cell_range[1:]
                # linha += 1
                # continue
        elif len(str(cell_range)) > 5:
            error_PO += 1
            aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            # linha += 1
            # continue
        else:
            # linha += 1
            pass

        # linha = 2

        ### VALIDAÇÃO DO PART-NUMBER

        # while linha != qtd_linhas + 1:
        # print(linha)
        linha_validada += 1
        cell_range = aba[f"D{linha}"].value
        if bool(cell_range) is False:
            aba[f"D{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            error_PN += 1
            # linha += 1
            # continue
        elif "!" in cell_range or \
             "@" in cell_range or \
             "$" in cell_range or \
             "%" in cell_range or \
             "&" in cell_range or \
             "*" in cell_range or \
             "(" in cell_range or \
             ")" in cell_range or \
             "'" in cell_range or \
             ":" in cell_range:
            aba[f"D{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            error_PN += 1
            # linha += 1
            # continue
        elif "1P" in cell_range:
            aba[f"D{linha}"] = cell_range[2:]
            # linha += 1
            # continue
        elif "30P" in cell_range:
            aba[f"D{linha}"] = cell_range[3:]
            # linha += 1
            # continue
        else:
            # linha += 1
            pass

        # linha = 2

        ### VALIDAÇÃO RFID DO PRODUTO

        # while linha != qtd_linhas + 1:
        # print(linha)
        linha_validada += 1

        cell_range = aba[f"E{linha}"].value
        if bool(cell_range) is False:
            aba[f"E{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            error_RFID += 1
            # linha += 1
            # continue
        elif len(str(cell_range)) != 24:
            aba[f"E{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            error_RFID += 1
            # linha += 1
            # continue
        elif "E" != cell_range[0]:
            aba[f"E{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            error_RFID += 1
            # linha += 1
            # continue
        if cell_range in repeticao_RFID:
            aba[f'E{linha}'].fill = PatternFill(fill_type='solid', fgColor='38B9FF')
            repeticao_RFID[cell_range].fill = PatternFill(fill_type='solid', fgColor='38B9FF')
            error_RFID += 1
        else:
            repeticao_RFID[cell_range] = aba[f'E{linha}']
        if cell_range in dfRFID:
            aba[f'E{linha}'].fill = PatternFill(fill_type='solid', fgColor="9B9B9B")
            error_RFID += 1
        else:
            pass
        # temp_linha = 2
        # repticao = False
        # while temp_linha != qtd_linhas + 1:
        #     linha_validada += 1
        #     if temp_linha == linha:
        #         temp_linha += 1
        #         continue
        #     elif aba[f"E{temp_linha}"].value == cell_range:
        #         aba[f"E{temp_linha}"].fill = PatternFill(fill_type="solid", fgColor="38B9FF")
        #         temp_linha += 1
        #         repticao = True
        #         continue
        #     else:
        #         temp_linha += 1
        # if repticao is True:
        #     aba[f"E{linha}"].fill = PatternFill(fill_type="solid", fgColor="38B9FF")
        #     error_RFID += 1
        #     linha += 1
        #     continue
        # linha += 1

        # linha = 2

        ### VALIDAÇÃO DO SERIAL NUMBER

        # while linha != qtd_linhas + 1:
        # print(linha)
        linha_validada += 1
        cell_range = aba[f"F{linha}"].value
        if bool(cell_range) is False:
            aba[f"F{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            error_SN += 1
            # linha += 1
            # continue
        elif "S" == str(cell_range)[0]:
            aba[f"F{linha}"] = cell_range[1:]
        if "!" in str(cell_range) or \
            "@" in str(cell_range) or \
            "$" in str(cell_range) or \
            "%" in str(cell_range) or \
            "&" in str(cell_range) or \
            "*" in str(cell_range) or \
            "(" in str(cell_range) or \
            ")" in str(cell_range) or \
            "'" in str(cell_range) or \
            ":" in str(cell_range) or \
            "/" in str(cell_range):
            aba[f"F{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            error_SN += 1
            # linha += 1
            # continue
        if cell_range in repeticao_SN:
            aba[f'F{linha}'].fill = PatternFill(fill_type='solid', fgColor='38B9FF')
            repeticao_SN[cell_range].fill = PatternFill(fill_type='solid', fgColor='38B9FF')
            error_SN += 1
        else:
            repeticao_SN[cell_range] = aba[f'F{linha}']
        if cell_range in dfSerial:
            aba[f'F{linha}'].fill = PatternFill(fill_type='solid', fgColor='9B9B9B')
            error_SN += 1
        else:
            pass
        # temp_linha = 2
        # repticao = False
        # while temp_linha != qtd_linhas + 1:
        #     linha_validada += 1
        #     if temp_linha == linha:
        #         temp_linha += 1
        #         continue
        #     elif aba[f"F{temp_linha}"].value == cell_range:
        #         aba[f"F{temp_linha}"].fill = PatternFill(fill_type="solid", fgColor="38B9FF")
        #         temp_linha += 1
        #         repticao = True
        #         continue
        #     else:
        #         temp_linha += 1
        # if repticao is True:
        #     aba[f"F{linha}"].fill = PatternFill(fill_type="solid", fgColor="38B9FF")
        #     error_SN += 1
        #     linha += 1
        #     continue
        # linha += 1

        # linha = 2

        ### VALIDAÇÃO DA DATA

            ### Formatar a data, para ser copiada para as tabelas e estoque ###

        # while linha != qtd_linhas + 1:
        linha_validada += 1
        cell_range = aba[f"H{linha}"].value
        if bool(cell_range) is False:
            aba[f"H{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF7B00')
            error_Date += 1
            # linha += 1
            # continue
        elif type(cell_range) != datetime:
            aba[f"H{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF0000')
            error_Date += 1
            # linha += 1
            # continue
        elif cell_range > datetime.today():
            aba[f"H{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF0000')
            error_Date += 1
            # linha += 1
            # continue
        else:
            # linha += 1
            # continue
            pass
        # linha = 2

        ### VALIDAÇÃO USUÁRIO

        # while linha != qtd_linhas + 1:
        # linha_validada += 1
        # cell_range = aba[f"I{linha}"].value
        # if bool(cell_range) is False:
        #     aba[f"I{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF7B00')
        #     # error_usuario += 1
        #     # linha += 1
        #     # continue
        # else:
        #     # linha += 1
        #     # continue
        #     pass

        ### CHAVE DE RELACIONAMENTO

        aba[f'K{linha}'] = str(aba[f'E{linha}'].value) + str(aba[f'G{linha}'].value)

        if aba[f'K{linha}'].value in dfTblRec_ChaveRelacionamento:
            tem_df = dfTblRec.loc[dfTblRec['ChaveRelacionamento'] == aba[f'K{linha}'].value]
            for i, row in tem_df.iterrows():
                if row['DataEvidencia'] >= aba[f'H{linha}'].value:
                    aba[f'K{linha}'].fill = PatternFill(fill_type='solid', fgColor='E7E200')
                    error_ChaveRel += 1
                else:
                    continue
        else:
            pass

        ### LANÇAMENTO BANCO DE DADOS - DATA

        date = datetime.today()
        aba[f'L{linha}'] = date.strftime("%d/%m/%Y")

        ### LANÇAMENTO BANCO DE DADOS - USUÁRIO

        aba[f'M{linha}'] = 'Automatizado'

        linha += 1

    aba['A1'] = 'ChaveNF_Entrada'
    aba['B1'] = 'PedidoCompra'
    aba['C1'] = 'RFID_CxMaster/TagAtivo'
    aba['D1'] = 'PartNumber'
    aba['E1'] = 'RFID_Produto'
    aba['F1'] = 'SerialNumber'
    aba['G1'] = 'Local'
    aba['H1'] = 'DataEvidencia'
    aba['I1'] = 'Usuario(email)'
    aba['J1'] = 'ObsRecebimento'
    aba['K1'] = 'ChaveRelacionamento'
    aba['L1'] = 'LctoBD_Data'
    aba['M1'] = 'LctoBD_Usuario'

    print(f"Erro ChaveNF:{error_chave:>8}"
          f"\nErro Pedido: {error_PO:>8}"
          f"\nErro PartNumber: {error_PN:>4}"
          f"\nErro RFID: {error_RFID:>10}"
          f"\nErro SerialNumber:{error_SN:>3}"
          f"\nErro Data: {error_Date:>10}"
          f"\nErro Chave Relacionamento: {error_ChaveRel:>10}"
          f"\nCálulas validadas: {linha_validada:>4}")

    print(f'Tempo validação recebimento: {datetime.now() - tempo_recebimento}')
    print('Fim da validação')

    if error_chave > 0 or error_PO > 0 or error_PN > 0 \
       or error_RFID > 0 or error_SN > 0 or error_Date > 0 or error_ChaveRel:
        return 'Erro nos dados'
    else:
        return 'Sucesso'


def exp_validacao(aba, qtd_linhas):

    # Imports
    from openpyxl.styles import PatternFill
    from datetime import datetime
    import pandas as pd

    # Variáveis

    linha_validada = 0
    linha = 2
    error_chave = 0
    error_RFID = 0
    error_Date = 0
    # error_usuario = 0
    error_ChaveRel = 0

    # dfV17 = pd.read_excel(
    #     "C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\Gestão Estoque RFID - Estoque "
    #     "Consolidado V17 - Copia.xlsm", sheet_name="ItensArmazenados")
    dfTblExp = pd.read_excel(
        "C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\tblEvidenciaExpedicao.xlsm", sheet_name='Evidencias')
    # dfRFID = dfV17['Unnamed: 8'].tolist()
    # dfSerial = dfV17['Unnamed: 9'].tolist()
    dfTblExp_ChaveRelacionamento = dfTblExp['ChaveRelacionamento'].tolist()

    print('Início da validação - Expedição')

    while linha != qtd_linhas + 1:

        ### VALIDAÇÃO RFID DO PRODUTO

        linha_validada += 1

        cell_range = aba[f"A{linha}"].value
        if bool(cell_range) is False:
            aba[f"A{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            error_RFID += 1
        elif len(str(cell_range)) != 24:
            aba[f"A{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            error_RFID += 1
        elif "E" != cell_range[0]:
            aba[f"A{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            error_RFID += 1
        if cell_range in repeticao_RFID:
            aba[f'A{linha}'].fill = PatternFill(fill_type='solid', fgColor='38B9FF')
            repeticao_RFID[cell_range].fill = PatternFill(fill_type='solid', fgColor='38B9FF')
            error_RFID += 1
        else:
            repeticao_RFID[cell_range] = aba[f'A{linha}']

        ### VALIDAÇÃO DA CHAVE DE NOTA FISCAL

        linha_validada += 1
        cell_range = aba[f"B{linha}"].value
        if bool(cell_range) is False:
            aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF7B00")
            error_chave += 1
        elif type(cell_range) != int and cell_range.isnumeric() is False:
            aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            error_chave += 1
        elif len(cell_range) != 44:
            aba[f"B{linha}"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            error_chave += 1
        elif len(cell_range) == 44:
            pass

        ### VALIDAÇÃO DA DATA

        linha_validada += 1
        cell_range = aba[f"E{linha}"].value
        if bool(cell_range) is False:
            aba[f"E{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF7B00')
            error_Date += 1
        elif type(cell_range) != datetime:
            aba[f"E{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF0000')
            error_Date += 1
        elif cell_range > datetime.today():
            aba[f"E{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF0000')
            error_Date += 1
        else:
            pass

        ### VALIDAÇÃO USUÁRIO

        # linha_validada += 1
        # cell_range = aba[f"F{linha}"].value
        # if bool(cell_range) is False:
        #     aba[f"F{linha}"].fill = PatternFill(fill_type='solid', fgColor='FF7B00')
        #     error_usuario += 1
        # else:
        #     pass

        ### CHAVE DE RELACIONAMENTO

        aba[f"H{linha}"] = str(aba[f"A{linha}"].value) + str(aba[f"D{linha}"].value)

        if aba[f"H{linha}"].value in dfTblExp_ChaveRelacionamento:
            tem_df = dfTblExp.loc[dfTblExp['ChaveRelacionamento'] == aba[f"H{linha}"].value]
            for i, row in tem_df.iterrows():
                if row['DataEvidencia'] >= aba[f"E{linha}"].value:
                    aba[f"H{linha}"].fill = PatternFill(fill_type='solid', fgColor='E7E200')
                    error_ChaveRel += 1
                else:
                    continue
        else:
            pass

        ### lANÇAMENTO BANCO DE DADOS - DATA

        date = datetime.today()
        aba[f"I{linha}"] = date.strftime("%d/%m/%Y")

        ### LANÇAMENTO BANCO DE DADOS - USUÁRIO

        aba[f"J{linha}"] = 'Automatizado'

        linha += 1

    aba[f'A1'] = 'RFID_Produto'
    aba[f'B1'] = 'ChaveNF_Saida'
    aba[f'C1'] = 'OrdemVenda'
    aba[f'D1'] = 'Local'
    aba[f'E1'] = 'DataEvidencia'
    aba[f'F1'] = 'Usuario(email)'
    aba[f'G1'] = 'ObsExpedicao'
    aba[f'H1'] = 'ChaveRelacionamento'
    aba[f'I1'] = 'LctoBD_Data'
    aba[f'J1'] = 'LctoBD_Usuario'

    print(f"Erro ChaveNF:{error_chave:>8}"
          f"\nErro RFID: {error_RFID:>10}"
          f"\nErro Data: {error_Date:>10}"
          f"\nErro Chave Relacionamento: {error_ChaveRel}"
          f"\nCálulas validadas: {linha_validada:>4}")

    print('Fim da validação')

    if error_RFID > 0 or error_chave > 0 or error_Date > 0 or error_ChaveRel > 0:
        return 'Erro nos dados'
    else:
        return 'Sucesso'
