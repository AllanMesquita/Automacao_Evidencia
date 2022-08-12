def popular_V17(aba, qtd_linhas):

    import openpyxl as xl
    import pandas as pd
    from openpyxl.styles import PatternFill, Font
    from datetime import datetime
    import xlwings

    tempo_popular = datetime.now()

    v17 = xl.load_workbook(
        "C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\"
        "Teste - Gestão Estoque RFID - Estoque Consolidado V17 - 14.01.2022 - Copia.xlsm", keep_vba=True)
    tbl_exp = xl.load_workbook("C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\tblEvidenciaExpedicao.xlsm", keep_vba=True)
    tbl_rec = xl.load_workbook(
        "C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\tblEvidenciaRecebimento.xlsm", keep_vba=True)
    df_mastersaf = pd.read_excel(
        "C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\2022 - NFs Entrada Mastersaf.xlsx", sheet_name='NFsEntradaItens')
    # df_exp = pd.read_excel("C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\tblEvidenciaExpedicao.xlsm", sheet_name='Evidencias')
    v17.active
    v17_sheets = v17.sheetnames
    aba_v17 = v17[v17_sheets[2]]
    tbl_rec.active
    tbl_rec_sheets = tbl_rec.sheetnames
    aba_tblRec = tbl_rec[tbl_rec_sheets[0]]
    tbl_exp.active
    tbl_exp_sheets = tbl_exp.sheetnames
    aba_tblExp = tbl_exp[tbl_exp_sheets[1]]

    colunas_v17 = ['A', 'B', 'H', 'K', 'I', 'J', 'C', 'E', 'F', 'G', 'D']
    colunas_tblRec = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
    colunas_evid = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'J', 'K']
    colunas_tblExp = ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA',
                      'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI']
    colunas_v17_base = ['F', 'L', 'M', 'N', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB',
                       'AC', 'AD' ,'AE', 'AF', 'AG', 'AH']
    chaveRelac_dic = {}
    delete_row = []

    ultima_linha_tblRec = len(aba_tblRec['A']) + 1
    ultima_linha_v17 = len(aba_v17['A']) + 1
    backup_ultima_linha_v17 = ultima_linha_v17
    ultima_linha_tblExp = len(aba_tblExp['A']) + 1
    linha = 2
    linha_exp = 2
    qtd_linhsExp = len(aba_tblExp['H'])
    qtd_linhasV17 = len(aba_v17['G'])

    coluna_v17 = 0
    coluna_v17_base = 0
    coluna_tblRec = 0
    coluna_evid = 0
    coluna_tblExp = 0

    chave = ''
    pn = ''
    find_chave = ''

    data = datetime.today()
    data = data.strftime('%d/%m/%Y')

    font = Font(name='Arial', size=9)

    print('Início da atualização - Tbl.Recebimento')

    ### INSERÇÃO DA EVIDÊNCIA NA TBL RECEBIMENTO

    while linha != qtd_linhas + 1:
        cell_range = aba[f'A{linha}':f'M{linha}']
        for cell in cell_range:
            for data in cell:
                aba_tblRec[f'{colunas_tblRec[coluna_tblRec]}{ultima_linha_tblRec}'] = data.value
                coluna_tblRec += 1
        coluna_tblRec = 0
        linha += 1
        ultima_linha_tblRec += 1

    tbl_rec.save("C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\tblEvidenciaRecebimento.xlsm")
    linha = 2
    print('Fim da atualização - Tbl.Recebimento')

    ### INSERÇÃO DA EVIDÊNCIA NA PLANILHA ESTOQUE

    # while linha != qtd_linhas + 1:
    #     cell_range = aba[f'{colunas_evid[coluna_evid]}{linha}'].value
    #     aba_v17[f'{colunas_v17[coluna_v17]}{ultima_linha_v17}'] = cell_range
    #     coluna_v17 += 1
    #     linha += 1
    #     ultima_linha_v17 += 1

    while linha != qtd_linhas + 1:
        for col in colunas_v17:
            # for c in range(qtd_linhas - 1):
            if col == 'D':
                aba_v17[f'{col}{ultima_linha_v17}'] = 'Recebimento'
            else:
                cell_range = aba[f'{colunas_evid[coluna_evid]}{linha}'].value
                if colunas_evid[coluna_evid] == 'B':
                    aba_v17[f'{col}{ultima_linha_v17}'] = int(cell_range)
                else:
                    aba_v17[f'{col}{ultima_linha_v17}'] = cell_range
                    aba_v17[f'{col}{ultima_linha_v17}'].font = font
                if colunas_evid[coluna_evid] == 'A':
                    chave = cell_range
                    find_chave = df_mastersaf.loc[df_mastersaf['Chave de Acesso'] == chave]
                    # print(find_chave)
                    # print(find_chave.columns)
                    if find_chave.empty == True:
                        aba_v17[f'{col}{ultima_linha_v17}'].fill = PatternFill(fill_type='solid', fgColor='FF0000')
                    else:
                        find_org = find_chave['CNPJ/CPF do Destinatário']
                        if find_org.at[find_org.index[0]] == int('00447484000111'):
                            aba_v17[f'P{ultima_linha_v17}'] = 1
                            aba_v17[f'P{ultima_linha_v17}'].font = font
                        elif find_org.at[find_org.index[0]] == int('00447484000200'):
                            aba_v17[f'P{ultima_linha_v17}'] = 2
                            aba_v17[f'P{ultima_linha_v17}'].font = font
                        elif find_org.at[find_org.index[0]] == int('00447484000626'):
                            aba_v17[f'P{ultima_linha_v17}'] = 6
                            aba_v17[f'P{ultima_linha_v17}'].font = font
                        elif find_org.at[find_org.index[0]] == int('05437734000156'):
                            aba_v17[f'P{ultima_linha_v17}'] = 22
                            aba_v17[f'P{ultima_linha_v17}'].font = font
                        elif find_org.at[find_org.index[0]] == int('05437734000318'):
                            aba_v17[f'P{ultima_linha_v17}'] = 24
                            aba_v17[f'P{ultima_linha_v17}'].font = font
                        elif find_org.at[find_org.index[0]] == int('05437734000407'):
                            aba_v17[f'P{ultima_linha_v17}'] = 26
                            aba_v17[f'P{ultima_linha_v17}'].font = font
                        elif find_org.at[find_org.index[0]] == int('05437734000580'):
                            aba_v17[f'P{ultima_linha_v17}'] = 28
                            aba_v17[f'P{ultima_linha_v17}'].font = font
                        elif find_org.at[find_org.index[0]] == int('05437734000660'):
                            aba_v17[f'P{ultima_linha_v17}'] = 30
                            aba_v17[f'P{ultima_linha_v17}'].font = font
                        elif find_org.at[find_org.index[0]] == int('31546914000186'):
                            aba_v17[f'P{ultima_linha_v17}'] = 50
                            aba_v17[f'P{ultima_linha_v17}'].font = font
                elif colunas_evid[coluna_evid] == 'D':
                    pn = cell_range
                    find_chave = df_mastersaf.loc[df_mastersaf['Chave de Acesso'] == chave]
                    if find_chave.empty == True:
                        pass
                    else:
                        find_pn = find_chave.loc[find_chave['Cód. Produto'] == pn]
                        if find_pn.empty == True:
                            aba_v17[f'{col}{ultima_linha_v17}'].fill = PatternFill(fill_type='solid', fgColor='FF0000')
                        else:
                            find_valor = find_pn['Valor Unitário Comercial']
                            valor = find_valor.at[find_valor.index[0]].replace(',', '.')
                            aba_v17[f'AE{ultima_linha_v17}'] = float(valor)
                            aba_v17[f'AE{ultima_linha_v17}'].font = font
            coluna_evid += 1
        linha += 1
        ultima_linha_v17 += 1
        coluna_evid = 0
            #     linha += 1
            #     ultima_linha_v17 += 1
            # coluna_evid += 1
            # # coluna_v17 += 1
            # linha = 2
            # ultima_linha_v17 = backup_ultima_linha_v17

    ### AJUSTE DA TABELA

    aba_v17.tables['ItensArmazenados'].ref = f'A2:AT{len(aba_v17["A"])}'
    # aba_v17[f'A{backup_ultima_linha_v17}:AT{len(aba_v17["A"])}'].font = font
    # aba_v17[f'AE{len(aba_v17["A"])}'].number_format = 'Number'

    ### VALIDAÇÃO E RETIRADAS DOS ITENS DO ESTOQUE

    linha = 3

    while linha_exp != qtd_linhsExp + 1:
        chaveRelac_dic[aba_tblExp[f'H{linha_exp}'].value] = linha_exp
        linha_exp += 1

    while linha != qtd_linhasV17 + 1:
        cell_range = aba_v17[f'G{linha}'].value
        if cell_range in chaveRelac_dic:
            if aba_tblExp[f'E{chaveRelac_dic[cell_range]}'].value >= aba_v17[f'E{linha}'].value:
                for col in colunas_tblExp:
                    if col == 'AH':
                        aba_tblExp[f'AH{chaveRelac_dic[cell_range]}'] = 'data'
                    elif col == 'AI':
                        aba_tblExp[f'AI{chaveRelac_dic[cell_range]}'] = 'Automatizado'
                    else:
                        aba_tblExp[f'{col}{chaveRelac_dic[cell_range]}'] = aba_v17[f'{colunas_v17_base[coluna_v17_base]}{linha}'].value
                    coluna_v17_base += 1
                delete_row.append(linha)
                aba.delete_rows(linha, 1)
        linha += 1
        coluna_v17_base = 0

    print(delete_row)

    tbl_exp.save("C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\tblEvidenciaExpedicao.xlsm")
    v17.save("C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\"
             "Teste - Gestão Estoque RFID - Estoque Consolidado V17 - 14.01.2022 - Copia.xlsm")

    ### EXCLUIR LINHA DA PLANILHA

    # if len(delete_row) > 0:
    #     print(len(delete_row))
    #     print(delete_row)
    #     app = xlwings.App()
    #     wb = xlwings.Book("C:\\Users\\Mesqu\\Documents\\Projects\\Projeto_Melhoria_Evidencias\\"
    #              "Teste - Gestão Estoque RFID - Estoque Consolidado V17 - 14.01.2022 - Copia.xlsm")
    #     sheet = wb.sheets
    #     aba = sheet[2]
    #
    #     for count in range(0, len(delete_row)):
    #         aba.range(f'{delete_row[count]}:{delete_row[count]}').api.Delete()
    #         print(count)
    #
    #     wb.save()
    #     wb.close()
    #     app.quit()

    # df_chaveRelac = df_exp['ChaveRelacionamento'].tolist()
    #
    # while linha != len(aba_v17['G']) + 1:
    #     if aba_v17[f'G{linha}'].value in df_chaveRelac:
    #         temp_df = df_exp.loc[df_exp['ChaveRelacionamento'] == aba_v17[f'G{linha}'].value]
    #         for i, row in temp_df.iterrows():
    #             if row['DataEvidencia'] >= aba_v17[f'E{linha}'].value:
    #                 index = temp_df.index
    #                 for col in colunas_tblExp:
    #                     if col == 'AH':
    #                         aba_tblExp[f'AH{index[0] + 2}'] = 'data'
    #                     elif col == 'AI':
    #                         aba_tblExp[f'AI{index[0] + 2}'] = 'Automatizado'
    #                     else:
    #                         aba_tblExp[f'{col}{index[0] + 2}'] = aba_v17[f'{colunas_v17_base[coluna_v17_base]}{linha}'].value
    #                     coluna_v17_base += 1
    #                 # coluna_v17_base = 0
    #     # retirar linha do estoque
    #     linha += 1
    #     coluna_v17_base = 0

    print(f'Tempo popular: {datetime.now() - tempo_popular}')


