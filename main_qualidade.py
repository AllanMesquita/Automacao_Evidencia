"""
    Main Script
"""

import openpyxl as xl
from datetime import datetime
from Modulos import validacao
from openpyxl.styles import PatternFill
import os
import sys
import win32com.client

now = datetime.now()

resultado = ''

# path = "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\" + sys.argv[1] + "\\"
# file_name = sys.argv[2]
#path = "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\Temp\\"
#file_name = "202203281100 35220305437734000156550010000133101100010619_Expedição.xlsx"
path = 'C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\01 Processamento\\'

files = os.listdir("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\01 Processamento\\")
len_pasta = len(files)
    
if len_pasta != 0:

    win32 = win32com.client.Dispatch('Excel.Application')
    win32.Visible = False
    tblPA = win32.Workbooks.Open("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")
    aba_tblPA = tblPA.Worksheets('tbl')

    verificar_status = False
    qtd_linhas_tblPA = aba_tblPA.UsedRange.Rows.Count
    
    for value in range(1, qtd_linhas_tblPA + 1):
        if str(aba_tblPA.Range(f'E{value}')) == "EmProcessamento":
            verificar_status = True
            continue
    
    if verificar_status is False:

        for file_name in files:

            qtd_linhas_tblPA = aba_tblPA.UsedRange.Rows.Count
            
            aba_tblPA.Range(f'A{qtd_linhas_tblPA + 1}').Value = file_name.split('_')[0]
            aba_tblPA.Range(f'B{qtd_linhas_tblPA + 1}').Value = 'QualidadeEvidencia'
            aba_tblPA.Range(f'C{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
            aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = 'EmProcessamento' 
            
            tblPA.Save()

            wb = xl.load_workbook(path + file_name)
            wb.active
            sheet = wb.sheetnames
            aba = wb[sheet[0]]

            verif_evid = aba["A2"].value  # Verificador do tipo de evidência


            # Variáveis


            type_evidencia = ""
            validacao_local = False
            qtd_linhas = len(aba["A"])
            resultado = ''

            if bool(verif_evid) is False:
                #print("Erro!")
                resultado = 'Erro nos dados'
            elif len(verif_evid) == 44:
                type_evidencia = "Recebimento"
                # verificador de repetição de caracteres
                chave_nf = str(verif_evid)
                for c in chave_nf:
                    if chave_nf.count(c) == 44:
                        resultado = 'Erro nos dados'
                        break
                    else:
                        continue
            elif len(verif_evid) == 24 and "E" in verif_evid:
                type_evidencia = "Expedição"
            elif len(verif_evid) != 44:
                #print("Erro!")
                resultado = 'Erro nos dados'

            if type_evidencia == 'Recebimento':
                if aba['G2'].value == 'TERCA VIX' or aba['G2'].value == 'JR SAO' or aba['G2'].value == 'JR RIO' or aba['G2'].value == 'AGS RIO' or aba['G2'].value == 'NEXUS SAO':
                    pass
                else:
                    aba['G2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                    validacao_local = True
                    resultado = 'Erro local'
            elif type_evidencia == 'Expedição':
                if aba['D2'].value == 'TERCA VIX' or aba['D2'].value == 'JR SAO' or aba['D2'].value == 'JR RIO' or aba['D2'].value == 'AGS RIO' or aba['D2'].value == 'NEXUS SAO':
                    pass
                else:
                    aba['D2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                    validacao_local = True
                    resultado = 'Erro local'

            inicio = now
            #print(now)

            if validacao_local == False:
                if type_evidencia == "Recebimento":
                    resultado = validacao.rec_validation(aba, qtd_linhas)
                if type_evidencia == "Expedição":
                    resultado = validacao.exp_validacao(aba, qtd_linhas)

            wb.save(path + file_name)

            if resultado == 'Erro nos dados' or resultado == 'Erro local':
                os.replace(path + file_name, "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\03 Erro\\" + file_name)
                # Verificar o caso se haver o mesmo arquivo no destino
            #elif resultado == 'Sucesso':
                #atualizar.popular_V17(aba, qtd_linhas)

            # if error_chave > 0:
            #     print("Houve erro na coluna Chave NF")
            # else:
            #     print("Coluna Chave NF validada")

            aba_tblPA.Range(f'D{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
            aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = resultado

            tblPA.Save()

        tblPA.Save()
        win32.Application.Quit()
    
    else:
        outlook = win32com.client.Dispatch("outlook.application")

        mail = outlook.CreateItem(0)

        mail.To = 'allan.mesquita@global.ntt'
        mail.Subject = 'Error-Log - main_qualidade.py - "EmProcessamento"'
        mail.HTMLBody = '<h3>This is HTML Body</h3>'
        mail.Body = f"""Houve um erro na verificação da qualidade da evidência.
        
Existe outro processo com o status "EmProcessamento".
	
Att.
	
Python"""

        mail.Send()

else:
    resultado = 'Sem arquivos na pasta.'
    print('teste')

#print(inicio)
print(resultado)
print(f'Tempo total: {datetime.now() - now}')
