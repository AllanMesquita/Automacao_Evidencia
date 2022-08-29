"""
    Main Script
"""

import openpyxl as xl
from datetime import datetime
from Modulos import validacao
from openpyxl.styles import PatternFill
import os
# import sys
import win32com.client
import logging
import psycopg2

now = datetime.now()

name_log = str('C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Script\\Error_Log\\main_qualidade\\Error_Log_' + datetime.strftime(datetime.today(), '%d-%m-%Y %H.%M') + '.txt')
error_log_registro = str(
    'C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Script\\Error_Log\\main_v2\\Error_Log_\\Error-log - Registro\\' + datetime.strftime(
        datetime.today(), '%d-%m-%Y %H.%M') + '.txt')

global query_name, nome_evidencia, var_linha, query_id
resultado = ''
id_arquivo = ''
id = ''

# path = "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\" + sys.argv[1] + "\\"
# file_name = sys.argv[2]
#path = "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\Temp\\"
#file_name = "202203281100 35220305437734000156550010000133101100010619_Expedição.xlsx"

# path = 'C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\01 Processamento\\'
path = 'C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\01 Processamento\\'  # Troca dos diretórios 08.06.2022

# files = os.listdir("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\01 Processamento\\")
files = os.listdir("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\01 Processamento\\")  # Troca dos diretórios 08.06.2022
len_pasta = len(files)

# win32 = win32com.client.Dispatch('Excel.Application')
# win32.Visible = False
# tblPA = win32.Workbooks.Open("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")
# tblPA = win32.Workbooks.Open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")  # Troca dos diretórios 08.06.2022
# aba_tblPA = tblPA.Worksheets('tbl')
# tblPA = xl.open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")  # Troca dos diretórios 08.06.2022
# tblPA.active
# tblPA_sheet = tblPA.sheetnames
# aba_tblPA = tblPA[tblPA_sheet[0]]

verificar_status = False
# qtd_linhas_tblPA = aba_tblPA.UsedRange.Rows.Count
# qtd_linhas_tblPA = len(aba_tblPA['A'])

# INSERT QUERY
# aba_tblPA.Range(f'A{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(), '%Y%m%d%H%M')
# aba_tblPA.Range(f'B{qtd_linhas_tblPA + 1}').Value = 'QualidadeEvidencia'
# aba_tblPA.Range(f'C{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
# aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.today(), '%Y%m%d%H%M')
# aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'QualidadeEvidencia'
# aba_tblPA[f'C{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
# tblPA.Save()

try:
    query_name = datetime.strftime(datetime.now(), '%Y%m%d%H%M')
    query_id = query_name
    con = psycopg2.connect(
        host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
        dbname="logistic-control",
        user="logisticpsqladmin@psql-itlatam-logisticcontrol",
        password="EsjHSrS69295NzHu342ap6P!N",
        sslmode="require"
    )
    cur = con.cursor()
    # Pesquisa 'EmProcessamento'
    cur.execute(f"SELECT status FROM material_management.mm_tbl_processamento_automacoes WHERE status = 'EmProcessamento'")
    retorno = cur.fetchall()
    if bool(retorno) is False:
        cur.execute(f'INSERT INTO material_management.mm_tbl_processamento_automacoes (id_tbl, query, processamento_inicio,'
                    f'status) VALUES (%s, %s, %s, %s)',
                    (
                     query_id,
                     'QualidadeEvidencia',
                     datetime.now(),
                     'EmProcessamento'
                     )
                    )
        con.commit()
        cur.execute(f"SELECT id FROM material_management.mm_tbl_processamento_automacoes WHERE id_tbl = '{query_id}'")
        retorno = cur.fetchall()
        for c in retorno:
            id = c[0]
        cur.close()
        con.close()
    else:
        verificar_status = True

except Exception as error:
    logging.basicConfig(filename=error_log_registro, filemode='w', format='%(asctime)s %(message)s')
    logging.critical(f'- {error}', exc_info=True)

    tblPA = xl.open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")
    tblPA.active
    tblPA_sheet = tblPA.sheetnames
    aba_tblPA = tblPA[tblPA_sheet[0]]
    qtd_linhas_tblPA = len(aba_tblPA['A'])
    var_linha = qtd_linhas_tblPA + 1
    aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = query_name
    aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'QualidadeEvidencia'
    aba_tblPA[f'C{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
    aba_tblPA[f'E{var_linha}'] = 'EmProcessamento'
    tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

finally:
    # for value in range(1, qtd_linhas_tblPA + 1):
    #     if str(aba_tblPA.Range(f'E{value}')) == "EmProcessamento":
    #     # if str(aba_tblPA[f'E{value}'].value) == 'EmProcessamento':
    #         verificar_status = True
    #         continue

    if verificar_status is False:
        # var_linha = qtd_linhas_tblPA + 1
        # aba_tblPA.Range(f'C{var_linha}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
        # aba_tblPA.Range(f'E{var_linha}').Value = 'EmProcessamento'
        # tblPA.Save()
        # aba_tblPA[f'C{var_linha}'] = datetime.strftime(datetime.today(), '%d/%m%Y %H:%M')
        # aba_tblPA[f'E{var_linha}'] = 'EmProcessamento'

        try:
            if len_pasta > 0:

                for file_name in files:

                    # qtd_linhas_tblPA = aba_tblPA.UsedRange.Rows.Count
                    # qtd_linhas_tblPA = len(aba_tblPA['A'])

                    nome_evidencia = file_name.split('_')

                    # aba_tblPA.Range(f'A{qtd_linhas_tblPA + 1}').Value = nome_evidencia[0] + '_' + nome_evidencia[1]
                    # aba_tblPA.Range(f'B{qtd_linhas_tblPA + 1}').Value = 'QualidadeEvidencia'
                    # aba_tblPA.Range(f'C{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                    # aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = 'EmProcessamento'
                    # aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = file_name.split('_')[0]
                    # aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'QualidadeEvidencia'
                    # aba_tblPA[f'C{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                    # aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = 'EmProcessamento'

                    # tblPA.Save()
                    # tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

                    # QUERY ARQUIVO
                    try:
                        con = psycopg2.connect(
                            host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
                            dbname="logistic-control",
                            user="logisticpsqladmin@psql-itlatam-logisticcontrol",
                            password="EsjHSrS69295NzHu342ap6P!N",
                            sslmode="require"
                        )
                        cur = con.cursor()
                        cur.execute(
                                    f'INSERT INTO material_management.mm_tbl_processamento_automacoes (id_tbl, query, '
                                    f'processamento_inicio, status) VALUES (%s, %s, %s, %s)',
                                    (
                                     nome_evidencia[0] + '_' + nome_evidencia[1],
                                     'QualidadeEvidencia',
                                     datetime.now(),
                                     'EmProcessamento'
                                    )
                                    )
                        con.commit()
                        cur.execute(
                                    f"SELECT id FROM material_management.mm_tbl_processamento_automacoes WHERE "
                                    f"id_tbl = '{nome_evidencia[0] + '_' + nome_evidencia[1]}'"
                                    )
                        retorno = cur.fetchall()
                        for c in retorno:
                            id_arquivo = c[0]
                        cur.close()
                        con.close()
                    except Exception as error:
                        logging.basicConfig(filename=error_log_registro, filemode='w', format='%(asctime)s %(message)s')
                        logging.critical(f'- {error}', exc_info=True)

                        tblPA = xl.open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")
                        tblPA.active
                        tblPA_sheet = tblPA.sheetnames
                        aba_tblPA = tblPA[tblPA_sheet[0]]
                        qtd_linhas_tblPA = len(aba_tblPA['A'])
                        var_linha = qtd_linhas_tblPA + 1

                        aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = nome_evidencia[0] + '_' + nome_evidencia[1]
                        aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'QualidadeEvidencia'
                        aba_tblPA[f'C{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                        aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = 'EmProcessamento'
                        tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

                    finally:
                        wb = xl.load_workbook(path + file_name)
                        wb.active
                        sheet = wb.sheetnames
                        aba = wb[sheet[0]]

                        verif_evid = str(aba["A2"].value)  # Verificador do tipo de evidência


                        # Variáveis

                        type_evidencia = ""
                        validacao_local = False
                        qtd_linhas = len(aba["A"])
                        resultado = ''

                        if bool(verif_evid) is False:
                            #print("Erro!")
                            aba['A2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                            resultado = 'Erro nos dados'
                        elif len(verif_evid) == 44:
                            type_evidencia = "Recebimento"
                            # verificador de repetição de caracteres
                            chave_nf = str(verif_evid)
                            for c in chave_nf:
                                if chave_nf.count(c) == 44:
                                    aba['A2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                                    resultado = 'Erro nos dados'
                                    break
                                else:
                                    continue
                        elif len(verif_evid) == 24 and "E" in verif_evid:
                            type_evidencia = "Expedição"
                        elif len(verif_evid) != 44:
                            #print("Erro!")
                            aba['A2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                            resultado = 'Erro nos dados'

                        if type_evidencia == 'Recebimento':
                            local = aba['G2'].value
                            if local.strip() == 'TERCA VIX' or local.strip() == 'JR SAO' or local.strip() == 'JR RIO' or local.strip() == 'AGS RIO' or local.strip() == 'NEXUS SAO':
                                pass
                            else:
                                aba['G2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                                validacao_local = True
                                resultado = 'Erro local'
                        elif type_evidencia == 'Expedição':
                            local = aba['D2'].value
                            if local.strip() == 'TERCA VIX' or local.strip() == 'JR SAO' or local.strip() == 'JR RIO' or local.strip() == 'AGS RIO' or local.strip() == 'NEXUS SAO':
                                pass
                            else:
                                aba['D2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
                                validacao_local = True
                                resultado = 'Erro local'

                        inicio = now
                        #print(now)

                        if validacao_local == False:
                            if type_evidencia == "Recebimento":
                                resultado = validacao.rec_validation(aba, qtd_linhas, nome_evidencia[0] + '_' + nome_evidencia[1])
                            if type_evidencia == "Expedição":
                                resultado = validacao.exp_validacao(aba, qtd_linhas, nome_evidencia[0] + '_' + nome_evidencia[1])

                        wb.save(path + file_name)

                        if resultado == 'Erro nos dados' or resultado == 'Erro local':
                            # os.replace(path + file_name, "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\03 Erro\\" + file_name)
                            os.replace(path + file_name, "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\03 Erro\\" + file_name)  # Troca dos diretórios 08.06.2022
                            # Verificar o caso se haver o mesmo arquivo no destino
                        #elif resultado == 'Sucesso':
                            #atualizar.popular_V17(aba, qtd_linhas)

                        # if error_chave > 0:
                        #     print("Houve erro na coluna Chave NF")
                        # else:
                        #     print("Coluna Chave NF validada")

                        # aba_tblPA.Range(f'D{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                        # aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = resultado
                        # aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                        # aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = resultado

                        # tblPA.Save()
                        # tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

                        # STATUS FIM ARQUIVO
                        try:
                            con = psycopg2.connect(
                                host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
                                dbname="logistic-control",
                                user="logisticpsqladmin@psql-itlatam-logisticcontrol",
                                password="EsjHSrS69295NzHu342ap6P!N",
                                sslmode="require"
                            )
                            cur = con.cursor()
                            cur.execute(
                                        f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                                        f"processamento_fim = '{datetime.now()}',"
                                        f"status = '{resultado}' "
                                        f"WHERE id = '{id_arquivo}'")
                            con.commit()
                            cur.close()
                            con.close()

                        except Exception as error:
                            logging.basicConfig(filename=error_log_registro, filemode='w', format='%(asctime)s %(message)s')
                            logging.critical(f'- {error}', exc_info=True)

                            tblPA = xl.open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")
                            tblPA.active
                            tblPA_sheet = tblPA.sheetnames
                            aba_tblPA = tblPA[tblPA_sheet[0]]
                            qtd_linhas_tblPA = len(aba_tblPA['A'])
                            var_linha = qtd_linhas_tblPA + 1
                            aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = nome_evidencia[0] + '_' + nome_evidencia[1]
                            aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'QualidadeEvidencia'
                            aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                            aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = resultado
                            tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

                # aba_tblPA.Range(f'D{var_linha}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
                # aba_tblPA.Range(f'E{var_linha}').Value = 'Sucesso'
                # tblPA.Save()
                # win32.Application.Quit()
                # aba_tblPA[f'D{var_linha}'] = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
                # aba_tblPA[f'E{var_linha}'] = 'Sucesso'

                # tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

                # STATUS FIM PROCESSO
                try:
                    con = psycopg2.connect(
                        host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
                        dbname="logistic-control",
                        user="logisticpsqladmin@psql-itlatam-logisticcontrol",
                        password="EsjHSrS69295NzHu342ap6P!N",
                        sslmode="require"
                    )
                    cur = con.cursor()
                    cur.execute(
                                f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                                f"processamento_fim = '{datetime.now()}', "
                                f"status = 'Sucesso' "
                                f"WHERE id = '{id}'"
                    )
                    con.commit()
                    cur.close()
                    con.close()

                except Exception as error:
                    logging.basicConfig(filename=error_log_registro, filemode='w', format='%(asctime)s %(message)s')
                    logging.critical(f'- {error}', exc_info=True)

                    tblPA = xl.open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")
                    tblPA.active
                    tblPA_sheet = tblPA.sheetnames
                    aba_tblPA = tblPA[tblPA_sheet[0]]
                    qtd_linhas_tblPA = len(aba_tblPA['A'])
                    var_linha = qtd_linhas_tblPA + 1
                    aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = nome_evidencia[0] + '_' + nome_evidencia[1]
                    aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'QualidadeEvidencia'
                    aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                    aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = resultado
                    tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

            else:
                resultado = 'Sem arquivos na pasta.'      ### Bloco movido para a linha 148 - 05/05/2022
                # print('teste')
                # aba_tblPA.Range(f'D{var_linha}').Value = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                # aba_tblPA.Range(f'E{var_linha}').Value = resultado
                # aba_tblPA[f'D{var_linha}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                # aba_tblPA[f'E{var_linha}'] = resultado

                # tblPA.Save()
                # win32.Application.Quit()
                # tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

                # STATUS SEM ARQUIVO NA PASTA
                try:
                    con = psycopg2.connect(
                        host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
                        dbname="logistic-control",
                        user="logisticpsqladmin@psql-itlatam-logisticcontrol",
                        password="EsjHSrS69295NzHu342ap6P!N",
                        sslmode="require"
                    )
                    cur = con.cursor()
                    cur.execute(
                                f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                                f"processamento_fim = '{datetime.now()}', "
                                f"status = '{resultado}' "
                                f"WHERE id = '{id}'"
                    )
                    con.commit()
                    cur.close()
                    con.close()
                except Exception as error:
                    logging.basicConfig(filename=error_log_registro, filemode='w', format='%(asctime)s %(message)s')
                    logging.critical(f'- {error}', exc_info=True)

                    tblPA = xl.open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")
                    tblPA.active
                    tblPA_sheet = tblPA.sheetnames
                    aba_tblPA = tblPA[tblPA_sheet[0]]
                    qtd_linhas_tblPA = len(aba_tblPA['A'])
                    var_linha = qtd_linhas_tblPA + 1
                    aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = query_name
                    aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'QualidadeEvidencia'
                    aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                    aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = resultado
                    tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

        except Exception as error:
            logging.basicConfig(filename=name_log, filemode='w', format='%(asctime)s %(message)s')
            logging.critical(f'- {error}', exc_info=True)

            # aba_tblPA.Range(f'D{var_linha}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
            # aba_tblPA.Range(f'E{var_linha}').Value = 'Error-log'
            # aba_tblPA[f'D{var_linha}'] = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
            # aba_tblPA[f'E{var_linha}'] = 'Error-log'

            # aba_tblPA.Range(f'D{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')   ### Bloco movido para a linha 152 - 05/05/2022
            # aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = 'Error-log'
            # tblPA.Save()
            # win32.Application.Quit()
            # aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
            # aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = 'Error-log'

            # tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

            # STATUS ERROR-LOG
            '''Update query_id'''
            try:
                con = psycopg2.connect(
                    host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
                    dbname="logistic-control",
                    user="logisticpsqladmin@psql-itlatam-logisticcontrol",
                    password="EsjHSrS69295NzHu342ap6P!N",
                    sslmode="require"
                )
                cur = con.cursor()
                cur.execute(
                            f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                            f"processamento_fim = '{datetime.now()}',"
                            f"status = 'Error-log' "
                            f"WHERE id = '{id}'"
                )
                con.commit()
                '''Update id_arquivo'''
                cur.execute(
                            f"UPDATE material_management.mm_tbl_processamento_automacoes SET "
                            f"processamento_fim = '{datetime.now()}',"
                            f"status = 'Error-log' "
                            f"WHERE id = '{id_arquivo}'"
                )
                con.commit()
                cur.close()
                con.close()

            except Exception as error:
                logging.basicConfig(filename=error_log_registro, filemode='w', format='%(asctime)s %(message)s')
                logging.critical(f'- {error}', exc_info=True)

                tblPA = xl.open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")
                tblPA.active
                tblPA_sheet = tblPA.sheetnames
                aba_tblPA = tblPA[tblPA_sheet[0]]
                qtd_linhas_tblPA = len(aba_tblPA['A'])
                var_linha = qtd_linhas_tblPA + 1
                aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = query_name
                aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'QualidadeEvidencia'
                aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = resultado
                aba_tblPA[f'A{qtd_linhas_tblPA + 2}'] = nome_evidencia[0] + '_' + nome_evidencia[1]
                aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'QualidadeEvidencia'
                aba_tblPA[f'D{qtd_linhas_tblPA + 2}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
                aba_tblPA[f'E{qtd_linhas_tblPA + 2}'] = resultado
                tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

    else:
        # aba_tblPA.Range(f'D{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
        # aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = 'Error - EmProcessamento'
        # tblPA.Save()
        # win32.Application.Quit()
        # aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')
        # aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = 'Error - Emprocessamento'

        # tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

        try:
            con = psycopg2.connect(
                host="psql-itlatam-logisticcontrol.postgres.database.azure.com",
                dbname="logistic-control",
                user="logisticpsqladmin@psql-itlatam-logisticcontrol",
                password="EsjHSrS69295NzHu342ap6P!N",
                sslmode="require"
            )
            cur = con.cursor()
            cur.execute(
                f"INSERT INTO material_management.mm_tbl_processamento_automacoes (id_tbl, query, processamento_inicio,"
                f" processamento_fim, status) VALUES(%s, %s, %s, %s, %s)",
                (
                    query_id,
                    'QualidadeEvidencia',
                    datetime.now(),
                    datetime.now(),
                    'Error - EmProcessamento'
                )
                )
            con.commit()
            cur.close()
            con.close()

        except Exception as error:
            logging.basicConfig(filename=error_log_registro, filemode='w', format='%(asctime)s %(message)s')
            logging.critical(f'- {error}', exc_info=True)

            tblPA = xl.open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")
            tblPA.active
            tblPA_sheet = tblPA.sheetnames
            aba_tblPA = tblPA[tblPA_sheet[0]]
            qtd_linhas_tblPA = len(aba_tblPA['A'])
            var_linha = qtd_linhas_tblPA + 1
            aba_tblPA[f'A{qtd_linhas_tblPA + 1}'] = query_name
            aba_tblPA[f'B{qtd_linhas_tblPA + 1}'] = 'QualidadeEvidencia'
            aba_tblPA[f'D{qtd_linhas_tblPA + 1}'] = datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')
            aba_tblPA[f'E{qtd_linhas_tblPA + 1}'] = 'Error - Emprocessamento'
            tblPA.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\002 Tabelas\\tblProcessamentoAutomacoes.xlsx")

        finally:
            '''
                ENVIO DE E-MAIL
            '''
            outlook = win32com.client.Dispatch("outlook.application")

            mail = outlook.CreateItem(0)

            mail.To = 'allan.mesquita@global.ntt'
            mail.Subject = 'Erro-Log - main_qualidade.py - "EmProcessamento"'
            mail.HTMLBody = '<h3>This is HTML Body</h3>'
            mail.Body = f"""Houve um erro na varificação de qualidade das evidências.
        Outro Processo encontra-se com o status "EmProcessamento".
        
        Att.
        
        Python"""

            mail.Send()

            print('Error - EmProcessamento')

                # except Exception as error:
                #     logging.basicConfig(filename=name_log, filemode='w', format='%(asctime)s %(message)s')
                #     logging.critical(f'- {error}', exc_info=True)
                #
                #     aba_tblPA.Range(f'D{qtd_linhas_tblPA + 1}').Value = datetime.strftime(datetime.today(), '%d/%m/%Y %H:%M')   ### Bloco movido para a linha 152 - 05/05/2022
                #     aba_tblPA.Range(f'E{qtd_linhas_tblPA + 1}').Value = 'Error-log'
                #     tblPA.Save()
                #     win32.Application.Quit()

            # else:
            #     resultado = 'Sem arquivos na pasta.'      ### Bloco movido para a linha 148 - 05/05/2022
            #     # print('teste')

            # except Exception as error:
            #     logging.basicConfig(filename=name_log, filemode='w', format='%(asctime)s %(message)s')
            #     logging.critical(f'- {error}', exc_info=True)

#print(inicio)
print(resultado)
print(f'Tempo total: {datetime.now() - now}')
