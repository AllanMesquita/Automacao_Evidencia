from fileinput import filename
import traceback

import dateutil.parser
import openpyxl as xl
from openpyxl.styles import NamedStyle
from datetime import datetime
# import win32com.client
import os

# tempo = datetime.now()

# xll = win32com.client.Dispatch("Excel.Application")

# wb = xll.Workbooks.Open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\tblEvidenciaRecebimento.xlsm")

# xll.Application.Quit()

# print(datetime.now() - tempo)

# files = os.listdir("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\01 Processamento\\")

# print(len(files))

# if len(files) == 0:
#     print(0)
# else:
#     print('Erro')

# wb = xl.load_workbook("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Teste - Gestão Estoque RFID - Estoque Consolidado V17 - 24.03.2022.xlsm", keep_vba = True, read_only= True)
# wb.active
# sheet = wb.sheetnames
# aba = wb[sheet[2]]

# linha = 0
# count = 0

# # if linha == 0:
# #     if bool(aba[f'A1'].value) is True:
# #         linha += 1
# # else:
# while bool(aba[f'A{linha + 1}'].value) is True:
#     # print(f'{linha} - {aba[f"A{linha}"].value}')
#     linha += 1
#     count += 1

# print(linha)
# print(count)
# print(bool(aba[f'A26'].value))
# wb.save("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\03 Erro\\127633.xlsx")
# URL TEAMS
# url_temas = "https://dimensiondata.webhook.office.com/webhookb2/40cd17bf-322f-4d9c-9510-c7da651936a5@e3cf3c98-a978-465f-8254-9d541eeea73c/IncomingWebhook/5632e2389ca744cfad4cead6e2214acf/38a8d1ab-a097-4abc-9664-0d2d1e0b6640"

# import pymsteams

# msg = pymsteams.connectorcard(f"{url_temas}")
# msg.text("Bot test. :)")
# msg.send()

# import logging

# val = 'texto'

# name_log = str('C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Script\\Error_Log\\Error_Log_' + datetime.strftime(datetime.today(), '%d-%m-%Y %H.%M') + '.txt')
# try:
#     if vall == "texto":
#         print(val)
#     else:
#         print('Erro')
# except Exception as error:
#     logging.basicConfig(filename=name_log, filemode='a', format='%(asctime)s %(message)s')
#     logging.critical(f' - {error}', exc_info=True)

# import pandas
# tempo = datetime.now()
# excel = pandas.read_excel("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Teste - Gestão Estoque RFID - Estoque Consolidado V17 - 24.03.2022.xlsm", sheet_name="ItensArmazenados")
# print(datetime.now() - tempo)

# import pyodbc

# conn = pyodbc.connect('Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Access\\Estoque.accdb;')
# cursor = conn.cursor()
# cursor.execute('select * from Estoque')

# for row in cursor.fetchall():
#     print (row)

# local_emprocessamento = 'Teste'

# outlook = win32com.client.Dispatch("outlook.application")

# mail = outlook.CreateItem(0)

# mail.To = 'allan.mesquita@global.ntt'
# mail.Subject = 'Teste'
# mail.HTMLBody = '<h3>This is HTML Body</h3>'
# mail.Body = f"""Houve um erro na atualização da Planilha de estoque.
# O processo {local_emprocessamento} encontra-se com o status "EmProcessamento".

# Att.

# Python"""
# mail.Send()


# elif type(aba[f'H{linha}'].value) == str:
#     aba_to_date = datetime.strptime(aba[f'H{linha}'].value, '%d/%m/%Y')
#     if aba_tblExp[f'E{chaveRelac_dic[aba[f"K{linha}"].value]}'].value == str:
#         tblExp_to_date = datetime.strptime(aba_tblExp[f'E{chaveRelac_dic[aba[f"K{linha}"].value]}'].value, '')
#         if aba_tblExp[f'E{chaveRelac_dic[aba[f"K{linha}"].value]}'].value >= aba_to_date:
#             linha += 1
#             continue

# chave_nf = '33333333333333333333333333333333333333333333'

# # print(chave_nf.find('5'))

# for c in str(chave_nf):
#     if chave_nf.count(c) == 44:
#         print('44 repetido')
#         break
#     else:
#         continue

# files = os.listdir("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\01 Processamento\\")

# print(files[0].split('_')[0])


# win32 = win32com.client.Dispatch('Excel.Application')
# # win32.Visible = False
# tblrec = win32.Workbooks.Open("C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\06 Lixeira\\Testes\\02 Tabela\\Controle_Status_V17 - Copy.xlsx")
# aba_tblrec = tblrec.Worksheets('Sheet1')
#
# aba_tblrec.Range('A2').Value = 'Fim'
#
# tblrec.Save()
# win32.Application.Quit()

import pandas as pd

# excel = pd.read_excel(
#     "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\INDICADORES\\Bases\\2022 á 2027 - NFs Saída Mastersaf.xlsx",
#     sheet_name='Dados dos Itens')
#
# # df = excel['Unnamed: 17'].tolist()
# df = excel.loc[excel['Unnamed: 17'] == '33220305437734000580550010000026611100018356']
# dff = df['Unnamed: 28']
# lista = []
#
# for item in dff:
#     lista.append(float(item.replace(',', '.')))
#
# print(sum(lista))

# dic = {'teste1': 1, 'teste2': 2}
#
# for item in dic:
#     print(item)
#
# print(dic['teste1'])
#
# val = "TSP25210157"
# # var2 = '1.000,00'
# # var3 = "5"
# # print(bool(var3))
# # var3 = var3 if bool(var3) else 00
# #
# # # print(var2.replace('.', '').replace(',', '.'))
# # print(var3)
# #
# # lista = ['s']
# print(bool(val))
#
# data = '01/04/2022 15:00:01'
# data_conv = ''
# data_conv = datetime.strptime(data, '%d/%m/%Y %H:%M:%S')
# print(datetime.strptime(data, '%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y'))
# print(data_conv)

# for c in range(2, 66):
#     print("%s,", end=' ')
#
# lista1 = ('numero_nfe', 'serie', 'tipo_nfe', 'id_destinatario', 'id_fornecedor', 'id_natureza', 'chave_acesso', 'situacao', 'descricao_retorno', 'data_hora_emissao', 'data_hora_saida', 'data_hora_autorizacao', 'protocolo_autorizacao', 'data_hora_cancelamento', 'protocolo_cancelamento', 'motivo_cancelamento', 'ciencia_manifestacao', 'data_hora_manifestacao', 'id_transportadora', 'tipo_frete', 'codigo_antt', 'quantidade', 'especie', 'marca', 'numeracao_volume', 'peso_liquido', 'peso_bruto', 'base_calculo_icms', 'total_icms', 'total_icms_deson', 'total_fcp', 'total_fcp_uf_dest', 'total_icms_uf_dest', 'total_icms_uf_remet', 'base_calculo_icms_st', 'total_icms_st', 'total_fcp_st', 'total_fcp_st_ret', 'total_produtos_servicos', 'total_frete', 'total_seguro', 'total_desconto', 'total_ii', 'total_ipi', 'total_ipi_devolvido', 'total_pis', 'total_cofins', 'total_outras_despesas', 'total_nfe', 'vl_aprox_tot_trib', 'placa', 'uf', 'informacoes_adicionais_fisco', 'informacoes_complementares', 'usuario_consulta', 'data_consulta', 'id_cfop')
# lista = ('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')
# print(len(lista1))
#
# val = '1150,00'
# print(val.replace('.', ''))
#
# val = "numero_nfe, serie, tipo_nfe, id_destinatario, id_fornecedor, id_natureza, chave_acesso, situacao, descricao_retorno, datetime.strptime(data_hora_emissao, '%d/%m/%Y %H:%M:%S').strftime('%m/%d/%Y'), datetime.strptime(data_hora_saida, '%d/%m/%Y %H:%M:%S').strftime('%m/%d/%Y')"
# print(val[0:107])

# data = '01/01/0101'
#
# # print(datetime.strptime(data, '%d/%m/%Y %H:%M:%S').strftime('%m/%d/%Y'))
# print(data[0:10])

# val = ''
#
# print(bool(val))
#
# lista = ['erro chave', 'erro po', 'erro pn', 'erro rfid', 'erro sn', 'erro data', 'erro tbl']
#
# for c in range(len(lista)):
#     if bool(val) is False:
#         val = lista[c]
#     else:
#         val = val + ' / ' + lista[c]
#     # print(lista[c], end=' / ')
#
# print(val)


# excel = pd.read_excel(
#     "C:\\Users\\allan.mesquita\\NTT\\@AM BR Services and Operations - Privado\\INDICADORES\\Bases\\2022 á 2027 - NFs Entrada Mastersaf.xlsx",
#     sheet_name='Dados dos Itens')
#
# df = excel['Unnamed: 17'].tolist()
# df2 = excel.loc[excel['Unnamed: 17'] == '53220617698486000186550010000006531000058097']
# org = df2['Unnamed: 27']
# org2 = org.at[org.index[0]]
# # dff = df['Unnamed: 28']
# # lista = []
# print(org2)
# # print(dff)
# for c in df2:
#     print(c)

import openpyxl
from openpyxl.worksheet.table import Table

# wb = openpyxl.open("C:\\Users\\allan.mesquita\\Downloads\\Copy of Gestão Estoque RFID - Estoque Consolidado V17.1 - 14.07.2022.xlsm", keep_vba=True)
# ws = wb.active
# sheet = wb.sheetnames
# aba = wb[sheet[2]]
#
# total_linhas = len(aba['A'])
#
# print(total_linhas)
#
# print(aba.tables.items())
#
# table = wb['ItensArmazenados']
#
# # ws.append(['Teste'])
# # for c in range(1, 10):
# table.append(['Teste1'])
# table['B32374'] = 'Teste2'
# # aba.append(['Teste2'])
# # table[f'A{total_linhas}'] = 'Teste'
#
# # tab = Table(displayName='ItensArmazenados', insertRow=True)
# # tab.insertRow = True
# # ws.add_table(tab)
# # ws.tables['ItensArmazenados'].ref = 'A2:AT32815'
# print(aba.tables.items())
# # ws.tables.add(tab)
# print(len(aba['A']))
# wb.save("C:\\Users\\allan.mesquita\\Downloads\\Copy of Gestão Estoque RFID - Estoque Consolidado V17.1 - 14.07.2022.xlsm")


# nome = 'TERCA VIX_202207182043 32220705437734000318550010000461891100027998_Expedição.xlsx'
#
# nome2 = nome.split('_')
# print(nome2[0] + '_' + nome2[1])

# import psycopg2
#
# con = psycopg2.connect(
#     host = "psql-itlatam-logisticcontrol.postgres.database.azure.com",
#     dbname = "logistic-control",
#     user = "logisticpsqladmin@psql-itlatam-logisticcontrol",
#     password = "EsjHSrS69295NzHu342ap6P!N",
#     sslmode = "require"
# )
#
# cur = con.cursor()
#
# # id = 153
# #
# # for c in range(id, 172):
# cur.execute(f'DELETE FROM public.erros_evidencias WHERE id = 172')
# con.commit()

# #
# id_tbl = 'NEXUS SAO_202207190842 3522074626677100039855002000002565194877931'
# query = 'QualidadeEvidencia'
# processamentoInicio = '19/07/2022 13:00'
# processamentoFim = "20/07/2022 19:15"
# status = 'Sucesso'
#
# cur.execute(
#             f'INSERT INTO material_management.mm_tbl_processamento_automacoes (id_tbl, query, processamento_inicio, processamento_fim, status) VALUES (%s, %s, %s, %s, &s)',
#             (id_tbl, query, datetime.datetime.strptime(processamentoInicio, '%d/%m/%Y %H:%M'), datetime.datetime.strptime(processamentoFim, '%d/%m/%Y %H:%M'), status))
# con.commit()
#
# cur.execute(f"SELECT * FROM material_management.mm_tbl_processamento_automacoes WHERE id_tbl = '{id_tbl}'")
#
# resultado = cur.fetchall()
#
# print(bool(resultado))
# #
# # print(datetime.strptime(processamento_inicio, '%d/%m/%Y %H:%M'))
# #
# cur.close()
#
# con.close()
# #
# req_bory = [{'Id': '202207192040 32220704626426000700550010000492721684973626', 'Query': 'QualidadeEvidencia', 'ProcessamentoInicio': '19/07/2022 20:41', 'Status': 'Sucesso'}]
#
# print(req_bory[0]['Id'])

# local = 'AGS RIO '
#
# print(local)
#
# dic = {'chave': 0}
#
# dic['chave'] += 1
#
# print(dic['chave'])

# wb = xl.load_workbook('C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\GESTÃO DE ESTOQUE\\100 BcoDados\\003 Evidencias\\01 Processamento\\TERCA VIX - Para Teste_202207131524 32220728268233000784550010001859401000182252_Recebimento.xlsx')
# wb.active
# sheet = wb.sheetnames
# aba = wb[sheet[0]]
#
# print(type(aba['H10'].value))
#
# if type(aba['H10'].value) is datetime:
#     print('è datetime')
# else:
#     print('Não é')

# dic = {'12345': {'Erro': 6}}
#
# for erro in dic.items():
#     print(erro)

# nome = 'TERCA VIX_202207182008 3222070462642600070055001000049272168497362_Recebimento.xlsx'
#
# nome2 = nome.split('_')[0] + '_' + nome.split('_')[1]
#
# data = '20220802'
#
# # print(datetime.strptime(nome2, '%Y%m%d').strftime('%d/%m/%Y'))
# # print(datetime.strftime(datetime.strptime(data, '%Y%m%d'), '%d/%m/%Y'))
# print(nome2)

import psycopg2
import time

# con = psycopg2.connect(
#     host = "psql-itlatam-logisticcontrol.postgres.database.azure.com",
#     dbname = "logistic-control",
#     user = "logisticpsqladmin@psql-itlatam-logisticcontrol",
#     password = "EsjHSrS69295NzHu342ap6P!N",
#     sslmode = "require"
# )
#
# tempo = datetime.now()
#
# cur = con.cursor()
#
# id_tbl = datetime.strftime(datetime.now(), '%Y%m%d%H%M')
# id_tbl2 = id_tbl
# query = 'QualidadeEvidencia'
# dta_inicio = datetime.now()
# status = 'EmProcessamento'

### INSERT
# cur.execute(f'INSERT INTO material_management.mm_tbl_processamento_automacoes (id_tbl, query, processamento_inicio, '
#             f'status) VALUES (%s, %s, %s, %s)',
#             (id_tbl2, query, dta_inicio, status))
#
# con.commit()
#
# ### SELECT
id = ''
# cur.execute(f"SELECT status FROM material_management.mm_tbl_processamento_automacoes WHERE status = 'EmProcessamento'")
#
# resultado = cur.fetchall()
# for c in resultado:
#     id = c[0]
# print(resultado)
# print(bool(resultado))

# time.sleep(60)

# ### INSERT DATA/HORA FIM
# cur.execute(f'INSERT INTO material_management.mm_tbl_processamento_automacoes (processamento_fim) VALUES (%s)',
#             (datetime.now()))
# con.commit()
### UPDATE STATUS
# cur.execute(f"UPDATE material_management.mm_tbl_processamento_automacoes SET status = 'Sucesso' WHERE id = '80'")
# con.commit()
# # cur.execute(f"UPDATE material_management.mm_tbl_processamento_automacoes SET processamento_fim = '{datetime.now()}' WHERE id = '{id}'")
# # con.commit()
#
# print(datetime.now() - tempo)
#
# cur.close()
# con.close()
#
# print(datetime.now())
# print(datetime.strftime(datetime.now(), '%Y%m%d%H%M'))
from dateutil.parser import parse
# data = '13/08/2022'

# print(datetime.strptime(data, '%d/%m/%Y'))
# print(datetime.strftime(parse(data), '%Y/%m/%d'))

# try:
#     parse(data)
#     print(parse(data))
#
#     var = parse(data)
#     # print(var.day)
#     if var.day <= 12:
#         print(
#             datetime.strptime(datetime.strftime(parse(data), "%m/%d/%Y"), "%d/%m/%Y")
#         )
#     else:
#         print(var.day)
#         print(var.strftime("%d/%m/%Y"))
#     # data2 = datetime.strftime(parse(data), "%m/%d/%Y")
#     # print(data2)
#     # print(datetime.strptime(data2, "%d/%m/%Y"))
#     # print(
#     #     datetime.strptime(datetime.strftime(parse(data), "%m/%d/%Y"), "%d/%m/%Y")
#     # )
# except:
#     print(parse(data))
#     print("erro")

# print(
#         datetime.strptime(datetime.strftime(parse(data), "%m/%d/%Y"), "%d/%m/%Y")
#     )

# tbl_rec = xl.open("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\tblEvidenciaRecebimento.xlsm", keep_vba=True)
# tbl_rec.active
# tbl_rec_sheets = tbl_rec.sheetnames
# aba_tblRec = tbl_rec[tbl_rec_sheets[0]]
#
# data = '01/01/2001'
#
# aba_tblRec['H328131'] = parse(data)
#
# tbl_rec.save("C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\tblEvidenciaRecebimento.xlsm")
#
# print(datetime.strptime(datetime.strftime(parse(data), "%d/%m/%Y"), "%d/%m/%Y"))
#
# con = psycopg2.connect(
#     host = "psql-itlatam-logisticcontrol.postgres.database.azure.com",
#     dbname = "logistic-control",
#     user = "logisticpsqladmin@psql-itlatam-logisticcontrol",
#     password = "EsjHSrS69295NzHu342ap6P!N",
#     sslmode = "require"
# )
# cursor = con.cursor()
#

# cursor.execute(f"SELECT chave_acesso FROM public.nf_entrada2 WHERE chave_acesso = '35220915494741000143550010000370271407130895'")
# resultado = cursor.fetchall()
# for lista in resultado:
#     for dado in lista:
#         print(dado)

# tempo = datetime.now()
#
# cur = con.cursor()
#
# id = 4470
# ## INSERT
# for c in range(2528, id):
#     print(c)
#     cur.execute(
#                 f"UPDATE material_management.error_recebimento SET "
#                 f"erro = 'Execução de testes' "
#                 f"WHERE id = '{c}'"
#     )
#
#     con.commit()
#
# cur.close()
# con.close()

# var = '100.00'
#
# print(var.replace('.', '').replace(',', '.'))

# try:
#     print(10/10)
# except:
#     print('Error')
# else:
#     print('else')
# finally:
#     print('Teste')
#
# data = '10/08/2022'
#
# print(parse(data))
#
# var = parse(data)
# print(var.day)
# if var.day <= 12:
#     print(
#         datetime.strptime(datetime.strftime(parse(data), "%m/%d/%Y"), "%d/%m/%Y")
#     )
# else:
#     print(var.day)
#     print(var.strftime("%d/%m/%Y"))
# import traceback
# try:
#     var = [1, 3]
#     print(var[3])
# except:
#     print(traceback.format_exc())
# try:
#     data = ''
#
#     print(parse(data))
# except dateutil.parser.ParserError:
# import json
#
# data = open("C:\\Users\\allan.mesquita\\Downloads\\teste.json")
#
# obj = json.load(data)
#
# for linha in obj:
#     print(linha)
# #
# # for c in range(0, 11):
# #     if c == 5:
#         continue
#     else:
#         print(c)
import traceback
#
# try:
#     var = 2/0
#     print(var)
# except:
#     print(f'{traceback.format_exc()}')
#
# lista = {}
#
# for c in lista:
#     print(c['Nome'])

# dic = {'chave1': 2,
#        'chave2': 2
#        }
#
# for chave in dic.items():
#     print(chave[0])
# var = '29220776535764001891550030000271041420221645'
# var = '00000000000000000000000000000000000000000001'
# print(var.count('0'))
# for c in var:
#     # print(c)
#     if var.count(c) == 44:
#         # aba['A2'].fill = PatternFill(fill_type="solid", fgColor="FF0000")
#         # resultado = 'Erro nos dados'
#         # break
#         print('igual')
#     else:
#         continue
# var = '12/04/2022'
# print(parse(var))
# data = parse(var)
# print(data.day)
# if data.day <= 12:
#     data = datetime.strptime(datetime.strftime(data, "%m/%d/%Y"), "%d/%m/%Y")
#     print(data)
# data = '14/10/2022 16:52:31'
# # print(str(parse(data)))
# num = 1
#
# while num != 2 + 1:
#     try:
#         var = 1/0
#         num += 1
#     except:
#         print('Teste1')
#     finally:
#         num += 1
#
# print('teste')
# df_nfEntrada = pd.read_excel(
#         "C:\\Users\\allan.mesquita\\OneDrive - NTT\\Privado\\INDICADORES\\Bases\\2022 á 2027 - Nfs Entrada Mastersaf.xlsx",
#         sheet_name='Dados dos Itens'
#     )
#
# df = df_nfEntrada.loc[df_nfEntrada]
# df2 = df['Unnamed: 28']
#
# print(df)

# import logging
#
# name_log = str(
#     'C:\\Users\\allan.mesquita\\OneDrive - NTT\\Documents\\Projetos\\Automacao_Evidencias\\Script\\Error_Log\\Error-log - BD\\Error_Log_' + datetime.strftime(
#         datetime.today(), '%d-%m-%Y %H.%M') + '.txt')
# try:
#     print(1/0)
# except:
#     logging.basicConfig(filename=name_log, format='%(asctime)s %(message)s', filemode='w')
#     logging.critical(f'{traceback.format_exc()}', exc_info=True)
#
# var = 'K12345'
# try:
#     var2 = int(var[1:])
#
#     print(var2)
# except:
#     print(traceback.format_exc())
# print('Teste')
# var = '12/10/1989'
#
# print(parse(var))
# var = str(111111)
#
# print(var.count(str(1)))
# print(str(2)*44)
# var = '3322042878710900013055001000002573177048042'
# peso = 2
# resul = 0
# print(str(var[:43]))
# for c in var[::-1]:
#     print(int(c) * peso)
#     mult = int(c) * peso
#     resul += mult
#     if peso == 9:
#         peso = 2
#     else:
#         peso += 1
#
# print(resul/11, ' ', resul % 11, '-', 11-(resul % 11))

# dict = [{'@odata.etag': '', 'ItemInternalId': 'a577fb96-edb3-4330-ab52-f39200bab473', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000314981', 'SerialNumber': 'E00000000000000000314981', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000314981TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': '2fd3dcae-3abd-4b0c-bbb8-a370cc05c85f', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000314982', 'SerialNumber': 'E00000000000000000314982', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000314982TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': '6771826c-7247-4f8b-94c5-adc8987cbf3f', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000314983', 'SerialNumber': 'E00000000000000000314983', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000314983TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': 'f3dd1f63-55c5-421d-b045-eb5ab7649af0', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000314984', 'SerialNumber': 'E00000000000000000314984', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000314984TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': '69320096-243f-44c1-a03d-c4b8d2a4442e', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000314985', 'SerialNumber': 'E00000000000000000314985', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000314985TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': '61c6d585-e8dc-4f5e-aa1c-b55d73c4338f', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000314986', 'SerialNumber': 'E00000000000000000314986', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000314986TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': '3ec4f654-958d-4acc-9406-56cc587bdd83', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000314987', 'SerialNumber': 'E00000000000000000314987', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000314987TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': '8b6518a0-77e2-42e8-9569-a237e91809b8', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000314988', 'SerialNumber': 'E00000000000000000314988', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000314988TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': 'f47d6baf-5005-4f56-81ef-d06201bf9ca0', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000314989', 'SerialNumber': 'E00000000000000000314989', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000314989TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': 'ba295b59-ae38-4032-8f0b-e2561d8c42ab', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000314996', 'SerialNumber': 'E00000000000000000314996', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000314996TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': 'fcfa0c43-52fe-4f1d-875f-c13042489a06', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000334226', 'SerialNumber': 'E00000000000000000334226', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000334226TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': '033d9bd6-8674-482c-ba98-3aa9690070c9', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000334227', 'SerialNumber': 'E00000000000000000334227', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000334227TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': '1aaec54c-0c2f-48a0-acac-9aa479f53ff7', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000334228', 'SerialNumber': 'E00000000000000000334228', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000334228TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': 'f1a5936d-1040-4a7f-9dfb-c5fb8d923c5d', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000334229', 'SerialNumber': 'E00000000000000000334229', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000334229TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': 'd5550209-62d9-4db1-8b46-c0c4ea822e60', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000334230', 'SerialNumber': 'E00000000000000000334230', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000334230TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': 'f4a3dac8-4d6c-43db-9d26-50392045ab96', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000334231', 'SerialNumber': 'E00000000000000000334231', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000334231TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': '29e013cf-084f-4b4c-bd58-b41c356128f1', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000334232', 'SerialNumber': 'E00000000000000000334232', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000334232TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': 'be15bde7-1e57-44b6-ba9c-0eafc6d3cb6a', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000334233', 'SerialNumber': 'E00000000000000000334233', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000334233TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': 'e2baed00-669e-49b0-96a6-5428460c1ed0', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000334234', 'SerialNumber': 'E00000000000000000334234', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000334234TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': '59c4f461-4b2d-4b49-9c17-f84cae1cc564', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000334262', 'SerialNumber': 'E00000000000000000334262', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000334262TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': '40811949-a96b-4f15-9769-29665fc6b074', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000334263', 'SerialNumber': 'E00000000000000000334263', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000334263TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}, {'@odata.etag': '', 'ItemInternalId': 'e988f8ac-1f7d-4f41-8bda-1e828396d56f', 'ChaveNF_Entrada': '33220428787109000130550010000025731770480423', 'PedidoCompra': '15845', 'RFID_CxMaster/TagAtivo': '', 'PartNumber': 'MATERIAL-SITE', 'RFID_Produto': 'E00000000000000000334264', 'SerialNumber': 'E00000000000000000334264', 'Local': 'TERCA VIX', 'DataEvidencia': '06/04/2022', 'Usuario(email)': '', 'ObsRecebimento': '', 'ChaveRelacionamento': 'E00000000000000000334264TERCA VIX', 'LctoBD_Data': '17/08/2022 11:53', 'LctoBD_Usuario': 'Automatizado'}]
#
# print(dict[1][1])
#
# host = "psql-itlatam-logisticcontrol.postgres.database.azure.com"
# dbname = "logistic-control"
# user = "logisticpsqladmin@psql-itlatam-logisticcontrol"
# password = "EsjHSrS69295NzHu342ap6P!N"
# sslmode = "require"
# # Construct connection string
# conn_string = "host={0} user={1} dbname={2} password={3} sslmode={4}".format(host, user, dbname,
#                                                                              password,
#                                                                              sslmode)
# conn = psycopg2.connect(conn_string)
# print("Connection established")
# cursor = conn.cursor()
#
# chave = '35220101771935000215550030015594161010253163'
#
# cursor.execute(f"SELECT destinatario FROM material_management.master_saf_entrada WHERE chave_acesso = '{chave}'")
# resultado = cursor.fetchall()
# cursor.close()
# conn.close()
# print(resultado)
# print(bool(resultado))
#
# host = "psql-itlatam-logisticcontrol.postgres.database.azure.com"
# dbname = "logistic-control"
# user = "logisticpsqladmin@psql-itlatam-logisticcontrol"
# password = "EsjHSrS69295NzHu342ap6P!N"
# sslmode = "require"
# # Construct connection string
# conn_string = "host={0} user={1} dbname={2} password={3} sslmode={4}".format(host, user, dbname,
#                                                                              password,
#                                                                              sslmode)
# conn = psycopg2.connect(conn_string)
# print("Connection established")
# cursor = conn.cursor()
#
# cursor.execute(f"SELECT cod_produto, valor_unitario FROM material_management.master_saf_entrada_itens WHERE chave_acesso = '{chave}'")
# resultado = cursor.fetchall()
# print(resultado)
# print(bool(resultado))
#
# cursor.close()
# conn.close()
#
# var = [('LIC00593569B', '$5,838.64'), ('LIC00563557B', '$180,003.32')]
# dic = {}
# lista = []
#
# for item in var:
#     print(item)
#     dic[item[0]] = item[1]
#     dic_copy = dic.copy()
#     lista.append(dic_copy)
#     dic.clear()
#
# print(lista)
# print(dic)
# for c in lista:
#     for d in c.keys():
#         if d == 'LIC00593569B':
#             print(d)
#             print(str(c['LIC00593569B'])[1:].replace(',', ''))
#         else:
#             continue

host = "psql-itlatam-logisticcontrol.postgres.database.azure.com"
dbname = "logistic-control"
user = "logisticpsqladmin@psql-itlatam-logisticcontrol"
password = "EsjHSrS69295NzHu342ap6P!N"
sslmode = 'require'
# Construct connection string
conn_string = "host={0} user={1} dbname={2} password={3} sslmode={4}".format(host, user, dbname, password,
                                                                             sslmode)
conn = psycopg2.connect(conn_string)
print("Connection established")
cursor = conn.cursor()
